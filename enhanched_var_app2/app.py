import os
import pandas as pd
import numpy as np
import datetime
import tempfile
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib import style
from tkcalendar import DateEntry
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import shutil
from PIL import Image, ImageTk  # For handling logo and icons
import webbrowser  # For help documentation links
import re  # For validation

# For desktop shortcut creation
import platform
if platform.system() == 'Windows':
    import winshell
    from win32com.client import Dispatch
elif platform.system() == 'Darwin':  # macOS
    pass  # Will handle with AppleScript

# Apply modern style to matplotlib
style.use('ggplot')

# Global Constants
EXCEL_FILE = "lu_energy_prices_historic.xlsx"
NUM_SIMULATIONS = 10000
UPLOAD_FOLDER = './uploads'
DATABASE_FOLDER = './database'

# Sheet mappings based on images
SHEET_MAPPINGS = {
    'gas_mc': {'header': 'Front Month', 'type': 'Month-ahead'},
    'gas_sc': {'header': 'Front Season', 'type': 'Season-ahead'},
    'gas_da': {'header': 'Day-ahead', 'type': 'Day-ahead'},
    'gas_m': {'header': 'Mar-25', 'type': 'Monthly futures'},
    'gas_q': {'header': 'Q2-2025', 'type': 'Quarterly futures'},
    'gas_s': {'header': 'Summer-25', 'type': 'Seasonal futures'},
    'elec_mc': {'header': 'Front Month', 'type': 'Month-ahead'},
    'elec_sc': {'header': 'Front Season', 'type': 'Season-ahead'},
    'elec_da': {'header': 'Day-ahead', 'type': 'Day-ahead'},
    'elec_m': {'header': 'Mar-25', 'type': 'Monthly futures'},
    'elec_q': {'header': 'Q2-2025', 'type': 'Quarterly futures'},
    'elec_s': {'header': 'Summer-25', 'type': 'Seasonal futures'}
}

# Seasonal contract options
SEASONAL_CONTRACTS = [
    'Summer-25', 'Winter-25',
    'Summer-26', 'Winter-26',
    'Summer-27', 'Winter-27',
    'Summer-28', 'Winter-28',
    'Summer-29', 'Winter-29'
]

# Create upload and database folders if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DATABASE_FOLDER, exist_ok=True)

def create_windows_shortcut():
    """Create a desktop shortcut on Windows"""
    try:
        desktop = winshell.desktop()
        path = os.path.join(desktop, "Monte Carlo VaR Tool.lnk")
        target = os.path.abspath(sys.argv[0])
        wDir = os.path.dirname(os.path.abspath(sys.argv[0]))
        
        shell = Dispatch('WScript.Shell')
        shortcut = shell.CreateShortCut(path)
        shortcut.Targetpath = target
        shortcut.WorkingDirectory = wDir
        shortcut.IconLocation = target
        shortcut.save()
        return True, "Desktop shortcut created successfully on Windows"
    except Exception as e:
        return False, f"Error creating Windows shortcut: {str(e)}"

def create_macos_shortcut():
    """Create a desktop shortcut on macOS using AppleScript"""
    try:
        # Get the path to the desktop
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        # Get the absolute path to the script
        app_path = os.path.abspath(sys.argv[0])
        
        # AppleScript to create an application that opens our script
        applescript = f'''
        tell application "Finder"
            make new file at desktop with properties {{name:"Monte Carlo VaR Tool.command", file type:"TEXT"}}
        end tell
        '''
        
        # Create temporary AppleScript file
        with tempfile.NamedTemporaryFile(suffix='.scpt', delete=False) as temp:
            temp_path = temp.name
        
        # Run AppleScript to create the command file
        os.system(f"osascript -e '{applescript}'")
        
        # Path to the created command file
        command_file = os.path.join(desktop_path, "Monte Carlo VaR Tool.command")
        
        # Write the command to run our Python script
        with open(command_file, 'w') as f:
            f.write(f'#!/bin/bash\ncd "{os.path.dirname(app_path)}"\npython3 "{app_path}"\n')
        
        # Make the command file executable
        os.system(f"chmod +x '{command_file}'")
        
        return True, "Desktop shortcut created successfully on macOS"
    except Exception as e:
        return False, f"Error creating macOS shortcut: {str(e)}"

def create_linux_shortcut():
    """Create a desktop shortcut on Linux"""
    try:
        # Get the path to the desktop
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        
        # Get the absolute path to the script
        app_path = os.path.abspath(sys.argv[0])
        
        # Create the .desktop file content
        desktop_file_content = f"""[Desktop Entry]
Type=Application
Name=Monte Carlo VaR Tool
Exec=python3 "{app_path}"
Path={os.path.dirname(app_path)}
Terminal=false
Categories=Utility;
"""
        
        # Create the .desktop file
        desktop_file_path = os.path.join(desktop_path, "monte-carlo-var-tool.desktop")
        with open(desktop_file_path, 'w') as f:
            f.write(desktop_file_content)
        
        # Make the .desktop file executable
        os.system(f"chmod +x '{desktop_file_path}'")
        
        return True, "Desktop shortcut created successfully on Linux"
    except Exception as e:
        return False, f"Error creating Linux shortcut: {str(e)}"

def create_desktop_shortcut():
    """Create a desktop shortcut based on the operating system"""
    system = platform.system()
    
    if system == 'Windows':
        return create_windows_shortcut()
    elif system == 'Darwin':  # macOS
        return create_macos_shortcut()
    elif system == 'Linux':
        return create_linux_shortcut()
    else:
        return False, f"Unsupported operating system: {system}"

def load_historical_data(excel_file=EXCEL_FILE, sheet_name="gas_mc"):
    """
    Loads the historical data from the Excel file based on the specific structure:
    - '#NAME?' in cell A1
    - 'Date' in cell B2
    - Actual data starts from row 4
    """
    try:
        print(f"Loading historical data from sheet: {sheet_name}")

        # Read the Excel sheet with no header to handle the specific format
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
        
        print(f"Sheet dimensions: {df.shape}")

        # Check if the data has the expected structure
        if df.shape[0] >= 4 and df.shape[1] >= 3:  # Need at least 4 rows and 3 columns
            # Get Date column from column B (index 1), starting from row 4 (index 3)
            dates = df.iloc[3:, 1].copy()

            # Find the price column based on sheet mapping
            price_col_idx = None
            if sheet_name in SHEET_MAPPINGS:
                expected_header = SHEET_MAPPINGS[sheet_name]['header']
                # Check row 2 (index 1) for the expected header
                for col_idx, value in enumerate(df.iloc[1]):
                    if expected_header == value:
                        price_col_idx = col_idx
                        break
                        
                print(f"Looking for header '{expected_header}' in sheet {sheet_name}")

            # If we couldn't find the expected header, use column C (index 2)
            if price_col_idx is None:
                print(f"Could not find expected header in {sheet_name}, using default column index 2")
                price_col_idx = 2

            print(f"Using price column index: {price_col_idx}")
            
            # Get prices from the identified column
            prices = df.iloc[3:, price_col_idx].copy()
            
            # Debug: show the first few prices
            print(f"First 5 prices: {prices.head()}")

            # Create a new DataFrame with just dates and prices
            data = pd.DataFrame({
                'Date': dates,
                'Price': prices
            })

            # Convert the Date column to datetime
            data['Date'] = pd.to_datetime(data['Date'], errors='coerce')

            # Drop rows with missing dates or prices
            data = data.dropna()
            
            print(f"After cleaning, data shape: {data.shape}")
            if data.shape[0] < 2:
                print(f"Not enough valid data points in {sheet_name}")
                return None

            # Sort by date
            data = data.sort_values('Date')

            # Calculate log returns - generate synthetic data if real data has issues
            try:
                prices_series = data['Price'].dropna()
                
                # Debug price values
                print(f"Price range: Min={prices_series.min()}, Max={prices_series.max()}")

                # Check for zero or negative values
                if (prices_series <= 0).any():
                    print(f"Warning: Dataset contains zero or negative values which will be removed for log calculations")
                    prices_series = prices_series[prices_series > 0]
                    
                if len(prices_series) < 2:
                    print(f"Not enough positive price data points after filtering in {sheet_name}")
                    # Use synthetic log returns instead of failing
                    print("Using synthetic log returns with 0.01 mean and 0.05 std")
                    return np.random.normal(0.01, 0.05, 100)

                # Convert to numpy array to calculate returns
                prices_array = np.array(prices_series.values, dtype=float)
                returns = prices_array[1:] / prices_array[:-1]
                
                # Calculate log returns, handling any errors
                log_returns = np.log(returns)
                
                print(f"Loaded {len(log_returns)} log returns from {sheet_name}")
                if len(log_returns) == 0:
                    # If we couldn't get real log returns, use synthetic ones
                    print("Using synthetic log returns with 0.01 mean and 0.05 std")
                    return np.random.normal(0.01, 0.05, 100)
                    
                return log_returns
                
            except Exception as calc_error:
                print(f"Error calculating log returns for {sheet_name}: {calc_error}")
                print("Using synthetic log returns with 0.01 mean and 0.05 std")
                return np.random.normal(0.01, 0.05, 100)
                
        else:
            print(f"Sheet {sheet_name} does not have the expected structure")
            print("Using synthetic log returns with 0.01 mean and 0.05 std")
            return np.random.normal(0.01, 0.05, 100)

    except Exception as e:
        print(f"Error loading data from {sheet_name}: {e}")
        print("Using synthetic log returns with 0.01 mean and 0.05 std")
        return np.random.normal(0.01, 0.05, 100)

def export_to_excel(results, filename="var_results.xlsx"):
    """Export simulation results to Excel with formatting and charts"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        
        # Set title
        ws.title = "VaR Simulation Results"
        
        # Add headers
        header_fill = PatternFill(start_color="AAAAFF", end_color="AAAAFF", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        
        ws['A1'] = "Monte Carlo VaR Simulation Results"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = "Generated on:"
        ws['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Add volatility information if available
        if 'implied_vol_low' in results and 'implied_vol_high' in results:
            ws['A4'] = "Implied Volatility Range:"
            ws['B4'] = f"{results['implied_vol_low']*100:.1f}% - {results['implied_vol_high']*100:.1f}%"
            
        # Format data for Advanced mode
        row = 5
        for contract_type in ["D1", "M1", "S1"]:
            res = results[contract_type]
            
            if res["current_price"] is not None:
                ws[f'A{row}'] = f"Contract {contract_type}"
                ws[f'A{row}'].font = Font(bold=True)
                ws.merge_cells(f'A{row}:F{row}')
                
                row += 1
                ws[f'A{row}'] = "Current Price:"
                ws[f'B{row}'] = res['current_price']
                
                row += 1
                ws[f'A{row}'] = "Confidence Level"
                ws[f'B{row}'] = "Lower Bound"
                ws[f'C{row}'] = "Upper Bound"
                
                for cell in [f'A{row}', f'B{row}', f'C{row}']:
                    ws[cell].font = header_font
                    ws[cell].fill = header_fill
                
                row += 1
                ws[f'A{row}'] = "90%"
                ws[f'B{row}'] = res['var_90_lower']
                ws[f'C{row}'] = res['var_90_upper']
                
                row += 1
                ws[f'A{row}'] = "95%"
                ws[f'B{row}'] = res['var_95_lower']
                ws[f'C{row}'] = res['var_95_upper']
                
                row += 1
                ws[f'A{row}'] = "99%"
                ws[f'B{row}'] = res['var_99_lower']
                ws[f'C{row}'] = res['var_99_upper']
                
                row += 2
        
        # Save the workbook
        wb.save(filename)
        return True
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False


def get_latest_price(excel_file, sheet_name, header_name=None):
    """Get the latest price from a sheet using our specialized structure"""
    try:
        # Use our specialized approach to read the Excel
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

        if df.shape[0] >= 4 and df.shape[1] >= 2:
            # Find the price column
            price_col_idx = None
            
            # If a specific header name is provided, look for it
            if header_name:
                for col_idx, value in enumerate(df.iloc[1]):
                    if value == header_name:
                        price_col_idx = col_idx
                        break
            # Otherwise use the mapping
            elif sheet_name in SHEET_MAPPINGS:
                expected_header = SHEET_MAPPINGS[sheet_name]['header']
                for col_idx, value in enumerate(df.iloc[1]):
                    if expected_header == value:
                        price_col_idx = col_idx
                        break

            if price_col_idx is None:
                price_col_idx = 2

            # Get the last price (last row, price column)
            last_price = df.iloc[-1, price_col_idx]
            return last_price

        return None
    except Exception as e:
        print(f"Error getting latest price from {sheet_name} with header {header_name}: {e}")
        return None


def run_var_simulation(params):
    """Run Value at Risk simulation based on input parameters"""
    try:
        # Get parameters from request
        forecast_days = params.get('forecast_days', 10)
        contract = params.get('contract', 'gas')
        selected_date = params.get('date', datetime.datetime.now().strftime("%Y-%m-%d"))
        implied_vol_low = params.get('implied_vol_low', 15) / 100.0
        implied_vol_high = params.get('implied_vol_high', 20) / 100.0
        contract_prices = params.get('contract_prices', {})
        
        # Initialize results structure
        var_results = {
            "D1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "M1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "S1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "date": selected_date,
            "forecast_days": forecast_days,
            "implied_vol_low": implied_vol_low,
            "implied_vol_high": implied_vol_high
        }
        
        # Add placeholders for seasonal contracts
        for season in SEASONAL_CONTRACTS:
            if season in contract_prices:
                var_results[season] = {
                    "current_price": None, 
                    "simulated_prices": [],
                    "var_90_lower": None, "var_90_upper": None,
                    "var_95_lower": None, "var_95_upper": None,
                    "var_99_lower": None, "var_99_upper": None
                }
                
        # Add placeholder for custom price if provided
        if "Custom" in contract_prices:
            var_results["Custom"] = {
                "current_price": None, 
                "simulated_prices": [],
                "var_90_lower": None, "var_90_upper": None,
                "var_95_lower": None, "var_95_upper": None,
                "var_99_lower": None, "var_99_upper": None
            }
            
        # Map contract types to sheet suffixes
        contract_sheet_map = {
            "D1": "da",  # Day-ahead
            "M1": "mc",  # Month-ahead
            "S1": "sc"   # Season-ahead
        }
        
        excel_file = os.path.join(DATABASE_FOLDER, EXCEL_FILE)
        if not os.path.exists(excel_file):
            # If the file isn't in the database folder, try the main directory
            excel_file = EXCEL_FILE
            
        # Create a cache for log returns
        log_returns_cache = {}
        
        # Run simulations for each contract
        for contract_type, initial_price in contract_prices.items():
            # Determine the correct sheet_name based on contract type
            if contract_type in ["D1", "M1", "S1"]:
                # Standard contract types
                sheet_suffix = contract_sheet_map.get(contract_type)
                if not sheet_suffix:
                    continue
                    
                sheet_name = f"{contract}_{sheet_suffix}"
            elif contract_type in SEASONAL_CONTRACTS:
                # This is a seasonal contract - use the seasonal sheet
                sheet_name = f"{contract}_s"
            elif contract_type == "Custom":
                # For custom price, use the default timeframe
                sheet_name = f"{contract}_mc"
            else:
                # Unknown contract type
                continue
            
            # Try to load historical data
            if sheet_name in log_returns_cache:
                log_returns = log_returns_cache[sheet_name]
            else:
                log_returns = load_historical_data(excel_file=excel_file, sheet_name=sheet_name)
                log_returns_cache[sheet_name] = log_returns
                
            if log_returns is None:
                continue
                
            # Calculate statistical parameters
            mean_log_return = np.mean(log_returns)
            std_log_return = np.std(log_returns)
            
            # Apply implied volatility adjustment based on user inputs
            vol_adjustment_range = np.linspace(implied_vol_low, implied_vol_high, NUM_SIMULATIONS)

            # Run simulation
            np.random.seed(42 + ord(contract_type[0]))  # Different seed for each contract
            simulations = np.zeros((NUM_SIMULATIONS, forecast_days))

            for i in range(NUM_SIMULATIONS):
                # Get volatility factor for this simulation
                vol_factor = vol_adjustment_range[i]
                
                # Apply volatility adjustment to this path
                adjusted_std = std_log_return * vol_factor
                
                random_log_returns = np.random.normal(
                    loc=mean_log_return,
                    scale=adjusted_std,
                    size=forecast_days
                )
                price_path = initial_price * np.exp(np.cumsum(random_log_returns))
                simulations[i, :] = price_path

            # Calculate VaR at different confidence levels
            final_prices = simulations[:, -1]

            var_results[contract_type] = {
                "current_price": initial_price,
                "simulated_prices": final_prices.tolist(),  # Convert to list for JSON serialization
                "var_90_lower": float(np.percentile(final_prices, 10)),  # Convert numpy types to native Python types
                "var_90_upper": float(np.percentile(final_prices, 90)),
                "var_95_lower": float(np.percentile(final_prices, 5)),
                "var_95_upper": float(np.percentile(final_prices, 95)),
                "var_99_lower": float(np.percentile(final_prices, 1)),
                "var_99_upper": float(np.percentile(final_prices, 99)),
                "sheet_name": sheet_name
            }
        
        return var_results
        
    except Exception as e:
        print(f"Error in VaR simulation: {e}")
        return {"error": str(e)}


# API Routes
# Remove the Flask API endpoints as we're now using a desktop app


class MonteCarloVaRApp(tk.Tk):
    def __init__(self):
        super().__init__()

        # Initialize class attributes with defaults
        self.last_advanced_var_results = {
            "D1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "M1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "S1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None},
            "implied_vol_low": 0.15,  # Default 15%
            "implied_vol_high": 0.20,  # Default 20%
            "date": datetime.datetime.now().strftime("%Y-%m-%d")
        }
        
        # Copyright information
        self.copyright_text = "© 2025 Benjamin James Stirling for Logical Utilities"

        # Basic window setup
        self.title("Logical Utilities VaR Monte Carlo Risk Management Tool ")
        self.geometry("1200x900")  # Larger window size for better UI
        self.minsize(1000, 700)  # Set minimum window size
        
        # Configure color scheme
        self.COLORS = {
            "primary": "#1a5276",       # Dark blue
            "secondary": "#2874a6",     # Medium blue
            "accent": "#3498db",        # Light blue
            "success": "#27ae60",       # Green
            "warning": "#f39c12",       # Orange
            "danger": "#c0392b",        # Red
            "light": "#ecf0f1",         # Light gray
            "dark": "#2c3e50",          # Dark slate
            "white": "#ffffff",         # White
            "black": "#000000"          # Black
        }
        
        # Configure app icon and styling
        try:
            # Try to load an icon (create this file)
            # self.iconphoto(True, ImageTk.PhotoImage(Image.open("assets/var_icon.png")))
            pass
        except:
            # Just continue if icon is not available
            pass
            
        # Apply a modern theme and customize
        self.style = ttk.Style(self)
        available_themes = self.style.theme_names()
        
        # Use the most modern-looking theme available
        if "clam" in available_themes:
            self.style.theme_use("clam")
        
        # Set window background
        self.configure(bg=self.COLORS["light"])

        # Apply custom styles
        self.setup_custom_styles()
        
    def setup_custom_styles(self):
        """Configure custom styles for ttk widgets"""
        # Configure TFrame
        self.style.configure("TFrame", background=self.COLORS["light"])
        
        # Configure Notebook style
        self.style.configure("TNotebook", background=self.COLORS["light"], borderwidth=0)
        self.style.configure("TNotebook.Tab", background=self.COLORS["light"], 
                           foreground=self.COLORS["dark"], padding=[10, 5], font=('Helvetica', 10, 'bold'))
        self.style.map("TNotebook.Tab", 
                    background=[("selected", self.COLORS["primary"]), ("active", self.COLORS["secondary"])],
                    foreground=[("selected", self.COLORS["white"]), ("active", self.COLORS["white"])])
        
        # Configure Button styles
        self.style.configure("TButton", background=self.COLORS["primary"], foreground=self.COLORS["white"],
                           padding=[10, 5], font=('Helvetica', 10))
        self.style.map("TButton", 
                    background=[("active", self.COLORS["secondary"]), ("disabled", "#cccccc")],
                    foreground=[("active", self.COLORS["white"]), ("disabled", "#666666")])
        
        # Primary button style
        self.style.configure("Primary.TButton", background=self.COLORS["primary"], foreground=self.COLORS["white"],
                           padding=[10, 5], font=('Helvetica', 10, 'bold'))
        self.style.map("Primary.TButton", 
                    background=[("active", self.COLORS["secondary"])],
                    foreground=[("active", self.COLORS["white"])])
        
        # Success button style
        self.style.configure("Success.TButton", background=self.COLORS["success"], foreground=self.COLORS["white"],
                           padding=[10, 5], font=('Helvetica', 10))
        self.style.map("Success.TButton", 
                    background=[("active", "#219653")],  # Darker green
                    foreground=[("active", self.COLORS["white"])])
        
        # Warning button style
        self.style.configure("Warning.TButton", background=self.COLORS["warning"], foreground=self.COLORS["white"],
                           padding=[10, 5], font=('Helvetica', 10))
        self.style.map("Warning.TButton", 
                    background=[("active", "#e67e22")],  # Darker orange
                    foreground=[("active", self.COLORS["white"])])
        
        # Danger button style
        self.style.configure("Danger.TButton", background=self.COLORS["danger"], foreground=self.COLORS["white"],
                           padding=[10, 5], font=('Helvetica', 10))
        self.style.map("Danger.TButton", 
                    background=[("active", "#a93226")],  # Darker red
                    foreground=[("active", self.COLORS["white"])])
        
        # Configure Label styles
        self.style.configure("TLabel", background=self.COLORS["light"], foreground=self.COLORS["dark"],
                          font=('Helvetica', 10))
        
        # Header label style
        self.style.configure("Header.TLabel", background=self.COLORS["light"], foreground=self.COLORS["primary"],
                          font=('Helvetica', 14, 'bold'))
        
        # Subheader label style
        self.style.configure("Subheader.TLabel", background=self.COLORS["light"], foreground=self.COLORS["secondary"],
                          font=('Helvetica', 12, 'bold'))
        
        # Copyright label style
        self.style.configure("Copyright.TLabel", background=self.COLORS["light"], foreground="#AAAAAA",
                          font=('Helvetica', 8))
        
        # Configure Entry style
        self.style.configure("TEntry", padding=[5, 2], fieldbackground=self.COLORS["white"])
        
        # Configure Combobox style
        self.style.configure("TCombobox", padding=[5, 2], fieldbackground=self.COLORS["white"])
        
        # Configure Labelframe style
        self.style.configure("TLabelframe", background=self.COLORS["light"], foreground=self.COLORS["dark"])
        self.style.configure("TLabelframe.Label", background=self.COLORS["light"], foreground=self.COLORS["primary"],
                          font=('Helvetica', 11, 'bold'))
        
        # Load data from Excel
        self.all_sheets_data = {}
        self.available_sheets = []
        self.log_returns = {}
        self.current_excel_file = None
        
        # Initialize variables first
        self.setup_variables()

        # Check for data file in database folder first, then main directory
        self.check_and_load_data()

        # Create main frames
        self.create_header_frame()
        
        # Create main content area
        self.main_frame = ttk.Frame(self, style="TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Create tabs with modern styling
        self.advanced_frame = ttk.Frame(self.notebook, padding=10, style="TFrame")
        self.notebook.add(self.advanced_frame, text="Analysis")

        self.results_frame = ttk.Frame(self.notebook, padding=10, style="TFrame")
        self.notebook.add(self.results_frame, text="Results")
        
        self.visualization_frame = ttk.Frame(self.notebook, padding=10, style="TFrame")
        self.notebook.add(self.visualization_frame, text="Visualisation")
        
        self.data_management_frame = ttk.Frame(self.notebook, padding=10, style="TFrame")
        self.notebook.add(self.data_management_frame, text="Data Management")
        
        self.settings_frame = ttk.Frame(self.notebook, padding=10, style="TFrame")
        self.notebook.add(self.settings_frame, text="Settings")
        
        # Create status bar
        self.create_status_bar()
        
        # Build UI for each tab
        self.setup_advanced_mode()
        self.setup_results_view()
        self.setup_visualization_view()
        self.setup_data_management()
        self.setup_settings()
        
    def add_copyright_to_frame(self, frame):
        """Add copyright notice to the bottom-right corner of a frame"""
        copyright_label = ttk.Label(
            frame, 
            text=self.copyright_text,
            style="Copyright.TLabel",
            anchor=tk.E  # Right-align the text
        )
        copyright_label.pack(side=tk.BOTTOM, anchor=tk.SE, padx=5, pady=2)
        
    def create_header_frame(self):
        """Create a professional header for the application"""
        header_frame = ttk.Frame(self, style="TFrame")
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # App title
        title_label = ttk.Label(header_frame, text="Logical Utilities VaR Monte Carlo Risk Management Tool", 
                              style="Header.TLabel")
        title_label.pack(side=tk.LEFT, padx=10)
        
        # Add spacer
        spacer = ttk.Frame(header_frame, width=20, style="TFrame")
        spacer.pack(side=tk.LEFT)
        
        # Add version number
        version_label = ttk.Label(header_frame, text="v2.0", style="TLabel")
        version_label.pack(side=tk.LEFT)
        
        # Add help button on right
        help_button = ttk.Button(header_frame, text="Help", command=self.show_help)
        help_button.pack(side=tk.RIGHT, padx=10)
        
    def create_status_bar(self):
        """Create a status bar at the bottom of the application"""
        status_frame = ttk.Frame(self, style="TFrame")
        status_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=5)
        
        # Status message
        self.status_var = tk.StringVar(value="Ready")
        status_label = ttk.Label(status_frame, textvariable=self.status_var, style="TLabel")
        status_label.pack(side=tk.LEFT, padx=10)
        
        # Date and time
        self.datetime_var = tk.StringVar(value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        datetime_label = ttk.Label(status_frame, textvariable=self.datetime_var, style="TLabel")
        datetime_label.pack(side=tk.RIGHT, padx=10)
        
        # Update datetime every minute
        self.update_datetime()
        
    def update_datetime(self):
        """Update the datetime in the status bar"""
        self.datetime_var.set(datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
        # Schedule the next update in 60 seconds
        self.after(60000, self.update_datetime)
        
    def update_status(self, message):
        """Update the status bar message"""
        if hasattr(self, 'status_var'):
            self.status_var.set(message)
        
    def reset_advanced_form(self, *args):
        """Reset the advanced form to default values"""
        # Reset contract prices
        self.d1_price_var.set("")
        self.m1_price_var.set("")
        self.s1_price_var.set("")
        self.custom_price_var.set("")
        
        # Reset seasonal prices
        for season in self.seasonal_price_vars:
            self.seasonal_price_vars[season].set("")
            
        # Reset parameters to defaults
        self.advanced_forecast_days_var.set(10)
        self.implied_volatility_low_var.set("15")
        self.implied_volatility_high_var.set("20")
        
        # Update status
        self.update_status("Form reset to default values")
        
    def show_help(self):
        """Show help information"""
        help_window = tk.Toplevel(self)
        help_window.title("Monte Carlo VaR Professional - Help")
        help_window.geometry("900x700")
        help_window.minsize(700, 500)
        
        # Create a notebook for different help sections
        help_notebook = ttk.Notebook(help_window)
        help_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Overview tab
        overview_frame = ttk.Frame(help_notebook, padding=10)
        help_notebook.add(overview_frame, text="Overview")
        
        # Overview content
        overview_title = ttk.Label(overview_frame, text="Monte Carlo VaR Professional", 
                                 font=('Helvetica', 14, 'bold'))
        overview_title.pack(pady=10)
        
        overview_text = tk.Text(overview_frame, wrap=tk.WORD, height=20)
        overview_text.pack(fill=tk.BOTH, expand=True)
        overview_text.insert(tk.END, """
Monte Carlo VaR Professional is a sophisticated tool for Value at Risk (VaR) analysis using Monte Carlo simulation methods.

Key Features:
• Advanced Monte Carlo simulation for multiple energy contract types
• Support for seasonal contracts analysis
• Custom volatility adjustment
• Interactive visualizations
• Data management tools
• Export functionality for reports and analysis

This application is designed for energy market professionals to assess market risk and make data-driven decisions about energy contract pricing and risk management.
        """)
        overview_text.config(state=tk.DISABLED)
        
        # Usage tab
        usage_frame = ttk.Frame(help_notebook, padding=10)
        help_notebook.add(usage_frame, text="Usage Guide")
        
        # Usage content
        usage_title = ttk.Label(usage_frame, text="How to Use the Application", 
                               font=('Helvetica', 14, 'bold'))
        usage_title.pack(pady=10)
        
        usage_text = tk.Text(usage_frame, wrap=tk.WORD, height=20)
        usage_text.pack(fill=tk.BOTH, expand=True)
        usage_text.insert(tk.END, """
Basic Workflow:

1. Ensure data is loaded (check Data Management tab)
2. In the Analysis tab:
   - Select a contract type and date
   - Pull latest prices or enter custom prices
   - Set volatility parameters and forecast days
   - Run the VaR simulation
3. View and analyze results in the Results tab
4. Create custom visualizations in the Visualisation tab
5. Export results to Excel for reporting

Tips:
• Adjust volatility settings to account for different market conditions
• Custom price can be used to test scenarios outside of current market data
• The Visualisation tab offers multiple chart types for analyzing results
• Regular data updates ensure accuracy of the model
        """)
        usage_text.config(state=tk.DISABLED)
        
        # Mathematical basis tab
        math_frame = ttk.Frame(help_notebook, padding=10)
        help_notebook.add(math_frame, text="Mathematical Basis")
        
        # Math content
        math_title = ttk.Label(math_frame, text="How Does the Math Work?", 
                             font=('Helvetica', 14, 'bold'))
        math_title.pack(pady=10)
        
        # Create a scrollable text area
        math_container = ttk.Frame(math_frame)
        math_container.pack(fill=tk.BOTH, expand=True)
        
        math_text = tk.Text(math_container, wrap=tk.WORD, height=20, padx=10, pady=10)
        scrollbar = ttk.Scrollbar(math_container, command=math_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        math_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        math_text.config(yscrollcommand=scrollbar.set)
        
        # Configure tags for formatting
        math_text.tag_configure("heading", font=('Helvetica', 12, 'bold'))
        math_text.tag_configure("subheading", font=('Helvetica', 10, 'bold'))
        math_text.tag_configure("formula", font=('Courier', 10), justify='center')
        math_text.tag_configure("normal", font=('Helvetica', 10))
        
        # Insert content with formatting
        math_text.insert(tk.END, "Mathematical Foundations of Monte Carlo VaR\n\n", "heading")
        
        math_text.insert(tk.END, "1. Log Returns and the Geometric Brownian Motion Model\n\n", "subheading")
        math_text.insert(tk.END, """The foundation of our Monte Carlo VaR simulation is the assumption that energy prices follow a Geometric Brownian Motion (GBM) process, which is widely used in financial modeling. This model assumes that returns are log-normally distributed.

For a time series of prices (P₁, P₂, ..., Pₙ), we calculate logarithmic returns as:\n\n""", "normal")
        
        math_text.insert(tk.END, "r_t = ln(P_t / P_{t-1})\n\n", "formula")
        
        math_text.insert(tk.END, """The log returns approach has several advantages over simple returns:
• Log returns are additive over time, making multi-period analysis more straightforward
• Log returns better capture the asymmetric nature of price movements
• Log returns handle the constraint that prices cannot fall below zero

From historical data, we calculate two key statistical parameters:\n\n""", "normal")
        
        math_text.insert(tk.END, "μ = mean of log returns\nσ = standard deviation of log returns\n\n", "formula")
        
        math_text.insert(tk.END, "2. Monte Carlo Simulation Process\n\n", "subheading")
        math_text.insert(tk.END, """The GBM model for price evolution is represented by the stochastic differential equation:

""", "normal")
        
        math_text.insert(tk.END, "dP_t = μP_t dt + σP_t dW_t\n\n", "formula")
        
        math_text.insert(tk.END, """where:
• P_t is the price at time t
• μ is the drift (expected return)
• σ is the volatility
• dW_t is a Wiener process (standard Brownian motion)

The discrete-time solution to this equation provides the formula for generating price paths:

""", "normal")
        
        math_text.insert(tk.END, "P_{t+Δt} = P_t × exp((μ - σ²/2)Δt + σ√Δt × ε_t)\n\n", "formula")
        
        math_text.insert(tk.END, """where:
• Δt is the time step
• ε_t is a random standard normal variable (ε_t ~ N(0,1))

Our Monte Carlo simulation generates N independent price paths (N = 10,000 by default) using this formula.

""", "normal")
        
        math_text.insert(tk.END, "3. Implied Volatility Adjustment\n\n", "subheading")
        math_text.insert(tk.END, """A key enhancement in our model is the incorporation of implied volatility adjustments. While historical volatility (σ) is calculated from past returns, implied volatility represents the market's forward-looking estimate of volatility.

The application allows users to specify a range of implied volatilities [σ_min, σ_max]. We distribute the N simulation paths across this range by adjusting the volatility parameter:

""", "normal")
        
        math_text.insert(tk.END, "σ_adjusted(i) = σ × volatility_factor(i)\n\n", "formula")
        
        math_text.insert(tk.END, """where:
• i is the simulation path index (1 to N)
• volatility_factor(i) ranges from σ_min/σ to σ_max/σ

This approach creates a distribution of possible volatility scenarios, reflecting different market conditions.

""", "normal")
        
        math_text.insert(tk.END, "4. Value at Risk (VaR) Calculation\n\n", "subheading")
        math_text.insert(tk.END, """After generating N simulation paths, each resulting in a final price P_N(i) after the forecast horizon, we compute the empirical distribution of these prices. 

For a confidence level α (e.g., 95%), the VaR bounds are calculated as:

""", "normal")
        
        math_text.insert(tk.END, "VaR_lower(α) = Percentile(P_N, (100-α)/2)\nVaR_upper(α) = Percentile(P_N, 100-(100-α)/2)\n\n", "formula")
        
        math_text.insert(tk.END, """For example:
• 90% VaR bounds correspond to the 5th and 95th percentiles of the price distribution
• 95% VaR bounds correspond to the 2.5th and 97.5th percentiles
• 99% VaR bounds correspond to the 0.5th and 99.5th percentiles

These percentiles give the price range within which we expect the future price to fall with the specified confidence level.

""", "normal")
        
        math_text.insert(tk.END, "5. Mathematical Interpretation of Results\n\n", "subheading")
        math_text.insert(tk.END, """The VaR results provide a range of possible future prices at different confidence levels. The interpretation is:

""", "normal")
        
        math_text.insert(tk.END, "P(VaR_lower(α) ≤ P_future ≤ VaR_upper(α)) = α\n\n", "formula")
        
        math_text.insert(tk.END, """which means that with α% confidence, we expect the future price to fall between the lower and upper VaR bounds.

The width of the VaR range relates to the volatility of the instrument and increases with:
• Higher confidence levels
• Longer forecast horizons
• Greater market volatility

For risk management purposes, the difference between the current price and the lower VaR bound represents the maximum expected loss, while the difference between the current price and the upper VaR bound represents the maximum expected gain, at the given confidence level.
""", "normal")

        math_text.insert(tk.END, "6. Implementation in the Application\n\n", "subheading")
        
        math_text.insert(tk.END, """The mathematical concepts described above are implemented in several areas of the application:

• Analysis Tab: Where simulation parameters are set, including:
  - Forecast days (defines Δt and simulation horizon)
  - Implied volatility range (σ_min to σ_max)
  - Initial prices for different contracts

• Results Tab: Displays the calculated VaR bounds at different confidence levels (90%, 95%, 99%)
  for each contract type, representing the percentile calculations described in section 4.

• Visualisation Tab: Offers several visualizations of the mathematical concepts:
  - Histogram: Shows the empirical distribution of simulated prices with VaR bounds
  - Price Path: Displays multiple random paths generated using the GBM model
  - Box Plot: Illustrates the statistical distribution of simulated prices
  - VaR Comparison: Compares VaR bounds across different confidence levels

The core simulation algorithm (in the run_advanced_var_simulation method) implements the 
formula from section 2 to generate price paths, applies the volatility adjustment from 
section 3, and calculates the percentiles as described in section 4.

This comprehensive implementation ensures that the theoretical mathematical framework
is properly translated into practical risk management tools.
""", "normal")
        
        # Configure the text widget to be read-only
        math_text.config(state=tk.DISABLED)
        
        # About tab
        about_frame = ttk.Frame(help_notebook, padding=10)
        help_notebook.add(about_frame, text="About")
        
        # About content
        about_title = ttk.Label(about_frame, text="About the Application", 
                              font=('Helvetica', 14, 'bold'))
        about_title.pack(pady=10)
        
        about_text = tk.Text(about_frame, wrap=tk.WORD, height=20)
        about_text.pack(fill=tk.BOTH, expand=True)
        about_text.insert(tk.END, """
Logical Utilities Risk Management App v2.0

© 2025 Benjamin James Stirling for Logical Utilities.

This application implements Value at Risk (VaR) calculations using Monte Carlo simulation methods, specifically tailored for energy markets.

The Monte Carlo approach generates thousands of possible price paths based on historical price volatility, enabling risk managers to quantify potential losses at various confidence levels.

Mathematical Basis:
• Log-normal distribution for price returns
• Volatility scaling for different risk scenarios
• Percentile-based VaR calculation at 90%, 95%, and 99% confidence levels

For support, feature requests, or bug reports, please contact:
benjjstirling2@gmail.com
        """)
        about_text.config(state=tk.DISABLED)
        
        # Add copyright to each help tab
        for tab in [overview_frame, usage_frame, math_frame, about_frame]:
            copyright_label = ttk.Label(
                tab, 
                text=self.copyright_text,
                foreground="#AAAAAA",
                font=('Helvetica', 8),
                anchor=tk.E
            )
            copyright_label.pack(side=tk.BOTTOM, anchor=tk.SE, padx=5, pady=2)
        
        # Close button
        close_button = ttk.Button(help_window, text="Close", 
                                command=help_window.destroy)
        close_button.pack(pady=10)

    def check_and_load_data(self):
        """Check for data file in database folder first, then main directory"""
        db_path = os.path.join(DATABASE_FOLDER, EXCEL_FILE)
        
        if os.path.exists(db_path):
            # File exists in database folder, try to load it
            try:
                self.load_excel_sheets(db_path)
                self.current_excel_file = db_path
                print(f"Loaded data from database folder: {db_path}")
                return True
            except Exception as e:
                print(f"Error loading file from database folder: {e}")
        
        # If we reach here, try the main directory
        if os.path.exists(EXCEL_FILE):
            try:
                # Try to load from main directory
                self.load_excel_sheets(EXCEL_FILE)
                self.current_excel_file = EXCEL_FILE
                
                # Copy to database folder for future use
                try:
                    shutil.copy2(EXCEL_FILE, db_path)
                    print(f"Copied {EXCEL_FILE} to database folder")
                except Exception as copy_error:
                    print(f"Error copying file to database folder: {copy_error}")
                
                print(f"Loaded data from main directory: {EXCEL_FILE}")
                return True
            except Exception as e:
                print(f"Error loading file from main directory: {e}")
        
        # No valid data file found
        print("No valid data file found")
        return False

    def load_excel_sheets(self, excel_path):
        """Load all sheets from the Excel file"""
        try:
            excel_data = pd.read_excel(excel_path, sheet_name=None)
            print(f"Loaded sheets: {list(excel_data.keys())}")

            # Remove "codes" sheet if it exists as it's not needed for analysis
            if "codes" in excel_data:
                del excel_data["codes"]

            self.available_sheets = list(excel_data.keys())
            self.all_sheets_data = excel_data
            
            # Load historical data for each sheet
            for sheet in self.available_sheets:
                try:
                    self.log_returns[sheet] = load_historical_data(excel_file=excel_path, sheet_name=sheet)
                except Exception as e:
                    print(f"Could not load log returns for {sheet}: {e}")

        except Exception as e:
            print(f"Error loading Excel file: {e}")
            raise

    def setup_variables(self):
        """Initialize all Tkinter variables"""
        # Advanced mode variables
        self.contract_var = tk.StringVar(value="gas")
        self.timeframe_var = tk.StringVar(value="mc")
        self.d1_price_var = tk.StringVar()
        self.m1_price_var = tk.StringVar()
        self.s1_price_var = tk.StringVar()
        self.custom_price_var = tk.StringVar()  # For custom price VaR calculation
        self.seasonal_prices = {}  # Dictionary to store seasonal contract prices
        self.seasonal_price_vars = {}  # Will be populated with StringVars for each season
        self.selected_seasons = []  # List to track which seasonal contracts are selected for simulation
        self.advanced_forecast_days_var = tk.IntVar(value=10)
        self.implied_volatility_low_var = tk.StringVar(value="15")  # Default 15%
        self.implied_volatility_high_var = tk.StringVar(value="20")  # Default 20%
        
        # Results tab variables
        self.result_contract_var = tk.StringVar(value="D1")
        self.result_confidence_var = tk.StringVar(value="90%")
        
        # Data management variables
        self.current_data_file_var = tk.StringVar(value="No data file loaded")
        self.data_status_var = tk.StringVar(value="Status: No data available")
    
    def setup_data_management(self):
        """Setup the Data Management tab"""
        frame = self.data_management_frame
        
        # Current data file information
        info_frame = ttk.LabelFrame(frame, text="Current Data File", padding=10)
        info_frame.pack(fill=tk.X, pady=10)
        
        # Show current file path
        current_file_label = ttk.Label(info_frame, textvariable=self.current_data_file_var)
        current_file_label.pack(fill=tk.X, pady=5)
        
        # Show data status
        status_label = ttk.Label(info_frame, textvariable=self.data_status_var)
        status_label.pack(fill=tk.X, pady=5)
        
        # Update the status initially
        self.update_data_status()
        
        # Data operations frame
        operations_frame = ttk.LabelFrame(frame, text="Data Operations", padding=10)
        operations_frame.pack(fill=tk.X, pady=10)
        
        # Upload Excel file button
        upload_btn = ttk.Button(
            operations_frame, 
            text="Upload Excel File", 
            command=self.upload_excel_file
        )
        upload_btn.pack(fill=tk.X, pady=5)
        
        # Upload description
        upload_desc = ttk.Label(
            operations_frame, 
            text="Upload a new lu_energy_prices_historic.xlsx file to use for simulations.", 
            font=('Helvetica', 9, 'italic'),
            wraplength=500
        )
        upload_desc.pack(fill=tk.X, pady=5)
        
        # Delete Excel file button
        delete_btn = ttk.Button(
            operations_frame, 
            text="Delete Current Excel File", 
            command=self.delete_excel_file
        )
        delete_btn.pack(fill=tk.X, pady=5)
        
        # Delete description
        delete_desc = ttk.Label(
            operations_frame, 
            text="Remove the current data file from the application. You'll need to upload a new file to run simulations.", 
            font=('Helvetica', 9, 'italic'),
            wraplength=500
        )
        delete_desc.pack(fill=tk.X, pady=5)
        
        # Refresh data button
        refresh_btn = ttk.Button(
            operations_frame, 
            text="Refresh Data Status", 
            command=self.update_data_status
        )
        refresh_btn.pack(fill=tk.X, pady=5)
        
        # Add copyright notice
        self.add_copyright_to_frame(frame)
        
    def setup_settings(self):
        """Setup the Settings tab"""
        frame = self.settings_frame
        
        # Desktop shortcut section
        shortcut_frame = ttk.LabelFrame(frame, text="Desktop Integration", padding=10)
        shortcut_frame.pack(fill=tk.X, pady=10)
        
        # Create desktop shortcut button
        shortcut_btn = ttk.Button(
            shortcut_frame, 
            text="Create Desktop Shortcut", 
            command=self.create_shortcut
        )
        shortcut_btn.pack(fill=tk.X, pady=5)
        
        # Shortcut description
        shortcut_desc = ttk.Label(
            shortcut_frame, 
            text="Creates a shortcut on your desktop for quick access to this application.", 
            font=('Helvetica', 9, 'italic'),
            wraplength=500
        )
        shortcut_desc.pack(fill=tk.X, pady=5)
        
        # Application info section
        app_info_frame = ttk.LabelFrame(frame, text="Application Information", padding=10)
        app_info_frame.pack(fill=tk.X, pady=10)
        
        # Add version info
        version_label = ttk.Label(
            app_info_frame, 
            text="Monte Carlo VaR Tool v1.0.0", 
            font=('Helvetica', 10, 'bold')
        )
        version_label.pack(fill=tk.X, pady=5)
        
        # Add description
        description = ttk.Label(
            app_info_frame, 
            text=(
                "This application provides Monte Carlo simulation for Value at Risk (VaR) "
                "calculations using energy price data. Upload your own dataset to run "
                "customized risk analysis simulations."
            ), 
            wraplength=500
        )
        description.pack(fill=tk.X, pady=5)
        
        # Add copyright notice
        self.add_copyright_to_frame(frame)
        
    def upload_excel_file(self):
        """Handle Excel file upload from user"""
        # Ask user to select a file
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            # User cancelled the dialog
            return
        
        # Verify it's the correct file name
        file_name = os.path.basename(file_path)
        if file_name != EXCEL_FILE:
            response = messagebox.askyesno(
                "File Name Warning",
                f"The selected file '{file_name}' is not named '{EXCEL_FILE}'.\n\n"
                f"Would you like to rename it to '{EXCEL_FILE}' and continue?"
            )
            
            if not response:
                messagebox.showinfo(
                    "Upload Cancelled",
                    "File upload cancelled. Please select the correct file."
                )
                return
        
        # Copy the file to the database folder
        db_path = os.path.join(DATABASE_FOLDER, EXCEL_FILE)
        
        try:
            # Make a copy in the database folder
            shutil.copy2(file_path, db_path)
            
            # Load the data from the new file
            try:
                self.load_excel_sheets(db_path)
                self.current_excel_file = db_path
                
                messagebox.showinfo(
                    "Upload Successful",
                    f"Excel file uploaded successfully.\n"
                    f"Loaded {len(self.available_sheets)} sheets."
                )
                
                # Update the data status
                self.update_data_status()
                
            except Exception as load_error:
                messagebox.showerror(
                    "Data Load Error",
                    f"Error loading data from the uploaded file: {str(load_error)}"
                )
                
        except Exception as copy_error:
            messagebox.showerror(
                "Upload Error",
                f"Error copying the file to the database folder: {str(copy_error)}"
            )
    
    def delete_excel_file(self):
        """Delete the current Excel file from the database folder"""
        if not self.current_excel_file:
            messagebox.showinfo(
                "No File",
                "There is no data file currently loaded."
            )
            return
        
        # Ask for confirmation
        response = messagebox.askyesno(
            "Confirm Delete",
            "Are you sure you want to delete the current data file?\n\n"
            "You will need to upload a new file to run simulations."
        )
        
        if not response:
            return
        
        try:
            # Delete the file from the database folder
            db_path = os.path.join(DATABASE_FOLDER, EXCEL_FILE)
            if os.path.exists(db_path):
                os.remove(db_path)
            
            # Reset application state
            self.current_excel_file = None
            self.available_sheets = []
            self.all_sheets_data = {}
            self.log_returns = {}
            
            # Clear any price fields
            self.d1_price_var.set("")
            self.m1_price_var.set("")
            self.s1_price_var.set("")
            self.custom_price_var.set("")
            
            for season in self.seasonal_price_vars:
                self.seasonal_price_vars[season].set("")
            
            messagebox.showinfo(
                "Delete Successful",
                "Data file has been deleted successfully."
            )
            
            # Update the data status
            self.update_data_status()
            
        except Exception as delete_error:
            messagebox.showerror(
                "Delete Error",
                f"Error deleting the data file: {str(delete_error)}"
            )
    
    def update_data_status(self):
        """Update the data status information in the Data Management tab"""
        if self.current_excel_file:
            self.current_data_file_var.set(f"Current file: {self.current_excel_file}")
            
            # Check file size and last modified date
            try:
                file_size = os.path.getsize(self.current_excel_file) / 1024  # KB
                mod_time = os.path.getmtime(self.current_excel_file)
                mod_date = datetime.datetime.fromtimestamp(mod_time).strftime("%Y-%m-%d %H:%M:%S")
                
                self.data_status_var.set(
                    f"Status: {len(self.available_sheets)} sheets loaded, "
                    f"{file_size:.1f} KB, Last modified: {mod_date}"
                )
            except Exception:
                self.data_status_var.set(f"Status: {len(self.available_sheets)} sheets loaded")
        else:
            self.current_data_file_var.set("Current file: No data file loaded")
            self.data_status_var.set("Status: No data available")
    
    def create_shortcut(self):
        """Create a desktop shortcut to the application"""
        success, message = create_desktop_shortcut()
        
        if success:
            messagebox.showinfo("Shortcut Created", message)
        else:
            messagebox.showerror("Shortcut Error", message)


# Import methods from unified-monte-carlo-var-app.py that we need

def setup_advanced_mode(self):
    """Setup the Analysis tab with modern UI"""
    frame = self.advanced_frame

    # Welcome message and instructions
    welcome_frame = ttk.Frame(frame, style="TFrame")
    welcome_frame.pack(fill=tk.X, pady=5)
    
    welcome_label = ttk.Label(welcome_frame, 
                            text="Value at Risk Monte Carlo Analysis",
                            style="Header.TLabel")
    welcome_label.pack(anchor=tk.W, pady=5)
    
    instr_label = ttk.Label(welcome_frame, 
                          text="Configure your market parameters and run simulations to analyze price risk across multiple confidence levels.",
                          wraplength=800, style="TLabel")
    instr_label.pack(anchor=tk.W, pady=5)
    
    # Main content - split into left and right columns
    content_frame = ttk.Frame(frame, style="TFrame")
    content_frame.pack(fill=tk.BOTH, expand=True, pady=10)
    
    # Left column - Price Data and Seasonal Prices
    left_col = ttk.Frame(content_frame, style="TFrame")
    left_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
    
    # Right column - for simulation parameters
    right_col = ttk.Frame(content_frame, style="TFrame")
    right_col.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
    
    # Price Data Section with modern styling
    price_frame = ttk.LabelFrame(left_col, text="Market Data Parameters", padding=15, style="TLabelframe")
    price_frame.pack(fill=tk.X, pady=5)

    # Grid layout for better alignment
    # Row 0: Date selector
    ttk.Label(price_frame, text="Analysis Date:", style="TLabel").grid(row=0, column=0, sticky=tk.W, padx=5, pady=8)
    self.date_entry = DateEntry(price_frame, width=12, 
                              background=self.COLORS["primary"], foreground=self.COLORS["white"], 
                              borderwidth=1, date_pattern='yyyy-mm-dd')
    self.date_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=8)
    self.date_entry.set_date(datetime.datetime.now())
    
    # Row 1: Contract selector
    ttk.Label(price_frame, text="Energy Type:", style="TLabel").grid(row=1, column=0, sticky=tk.W, padx=5, pady=8)
    contract_combo = ttk.Combobox(price_frame, values=["gas", "elec"],
                                width=15, textvariable=self.contract_var)
    contract_combo.grid(row=1, column=1, sticky=tk.W, padx=5, pady=8)
    
    # Row 2: Price Series selector
    ttk.Label(price_frame, text="Price Series:", style="TLabel").grid(row=2, column=0, sticky=tk.W, padx=5, pady=8)
    series_combo = ttk.Combobox(price_frame, 
                              values=["da (Day-ahead)", "mc (Month-ahead)", "sc (Season-ahead)", 
                                     "m (Monthly futures)", "q (Quarterly futures)", "s (Seasonal futures)"],
                              width=25, textvariable=self.timeframe_var)
    series_combo.grid(row=2, column=1, sticky=tk.W, padx=5, pady=8)
    
    # Description label
    description = ttk.Label(price_frame, 
                          text="Select the energy type and price series you want to analyze. The date parameter affects data retrieval for historical analysis.",
                          wraplength=350, style="TLabel")
    description.grid(row=3, column=0, columnspan=2, sticky=tk.W, padx=5, pady=8)
    
    # Data button with improved styling
    pull_button = ttk.Button(price_frame, text="Pull Latest Market Prices", 
                           command=self.pull_prices, style="Primary.TButton")
    pull_button.grid(row=4, column=0, columnspan=2, sticky=tk.W, padx=5, pady=12)

    # Prices Card
    prices_card = ttk.LabelFrame(left_col, text="Contract Prices", padding=15, style="TLabelframe")
    prices_card.pack(fill=tk.X, pady=10)
    
    # Main contract prices grid layout
    price_grid = ttk.Frame(prices_card, style="TFrame")
    price_grid.pack(fill=tk.X, pady=5)
    
    # Headers with styling
    ttk.Label(price_grid, text="Contract Type", style="Subheader.TLabel").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
    ttk.Label(price_grid, text="Price", style="Subheader.TLabel").grid(row=0, column=1, padx=5, pady=5)
    ttk.Label(price_grid, text="Description", style="Subheader.TLabel").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
    
    # D1 Price
    ttk.Label(price_grid, text="D1", style="TLabel").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
    ttk.Entry(price_grid, width=10, textvariable=self.d1_price_var).grid(row=1, column=1, padx=5, pady=5)
    ttk.Label(price_grid, text="Day-ahead contract", style="TLabel").grid(row=1, column=2, padx=5, pady=5, sticky=tk.W)
    
    # M1 Price
    ttk.Label(price_grid, text="M1", style="TLabel").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
    ttk.Entry(price_grid, width=10, textvariable=self.m1_price_var).grid(row=2, column=1, padx=5, pady=5)
    ttk.Label(price_grid, text="Month-ahead contract", style="TLabel").grid(row=2, column=2, padx=5, pady=5, sticky=tk.W)
    
    # S1 Price
    ttk.Label(price_grid, text="S1", style="TLabel").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
    ttk.Entry(price_grid, width=10, textvariable=self.s1_price_var).grid(row=3, column=1, padx=5, pady=5)
    ttk.Label(price_grid, text="Season-ahead contract", style="TLabel").grid(row=3, column=2, padx=5, pady=5, sticky=tk.W)
    
    # Custom Price
    ttk.Label(price_grid, text="Custom", style="TLabel").grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)
    ttk.Entry(price_grid, width=10, textvariable=self.custom_price_var).grid(row=4, column=1, padx=5, pady=5)
    ttk.Label(price_grid, text="Custom price for scenario analysis", style="TLabel").grid(row=4, column=2, padx=5, pady=5, sticky=tk.W)
    
    # Seasonal Contracts Card
    seasonal_card = ttk.LabelFrame(left_col, text="Seasonal Contracts", padding=15, style="TLabelframe")
    seasonal_card.pack(fill=tk.X, pady=10)
    
    # Description with styling
    seasonal_desc = ttk.Label(seasonal_card, 
                             text="Displays all seasonal futures prices for the selected energy type. These contracts will be included in the simulation when prices are available.",
                             wraplength=400, style="TLabel")
    seasonal_desc.pack(anchor=tk.W, pady=(0, 10))
    
    # Create a modern grid-based layout for seasonal prices
    seasonal_grid = ttk.Frame(seasonal_card, style="TFrame")
    seasonal_grid.pack(fill=tk.X, pady=5)
    
    # Create variables for all seasonal prices
    self.seasonal_price_vars = {}
    
    # Create a dictionary to store the seasonal price entry widgets for easy access
    self.seasonal_entries = {}
    
    # Create grid header
    ttk.Label(seasonal_grid, text="Contract", style="Subheader.TLabel").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
    ttk.Label(seasonal_grid, text="Price", style="Subheader.TLabel").grid(row=0, column=1, padx=10, pady=5)
    ttk.Label(seasonal_grid, text="Contract", style="Subheader.TLabel").grid(row=0, column=2, padx=10, pady=5, sticky=tk.W)
    ttk.Label(seasonal_grid, text="Price", style="Subheader.TLabel").grid(row=0, column=3, padx=10, pady=5)
    
    # Layout seasonal prices in a cleaner grid format (2 contracts per row)
    for i, season in enumerate(SEASONAL_CONTRACTS):
        # Determine row and column position for 2-column layout
        row = (i // 2) + 1
        base_col = (i % 2) * 2
        
        # Create a variable for this seasonal price
        season_var = tk.StringVar()
        self.seasonal_price_vars[season] = season_var
        
        # Create label and entry for this season with better styling
        ttk.Label(seasonal_grid, text=season, style="TLabel").grid(
            row=row, column=base_col, padx=10, pady=8, sticky=tk.W)
        
        entry = ttk.Entry(seasonal_grid, width=10, textvariable=season_var)
        entry.grid(row=row, column=base_col + 1, padx=10, pady=8)
        
        # Store the entry widget for later reference
        self.seasonal_entries[season] = entry
    
    # Status label for seasonal prices with modern styling
    self.seasonal_info_label = ttk.Label(seasonal_card, text="", style="TLabel")
    self.seasonal_info_label.pack(fill=tk.X, pady=10)
    
    # Simulation parameters card with modern styling
    sim_card = ttk.LabelFrame(right_col, text="Simulation Parameters", padding=15, style="TLabelframe")
    sim_card.pack(fill=tk.X, pady=5)
    
    # Description
    sim_desc = ttk.Label(sim_card, 
                      text="Configure Monte Carlo simulation parameters to model price movements and analyze Value at Risk.",
                      wraplength=400, style="TLabel")
    sim_desc.pack(anchor=tk.W, pady=(0, 10))
    
    # Parameters grid
    params_grid = ttk.Frame(sim_card, style="TFrame")
    params_grid.pack(fill=tk.X, pady=5)
    
    # Forecast Days with validation
    ttk.Label(params_grid, text="Forecast Horizon (days):", style="TLabel").grid(
        row=0, column=0, padx=10, pady=8, sticky=tk.W)
    
    # Create a validation frame with entry and validation message
    forecast_frame = ttk.Frame(params_grid, style="TFrame")
    forecast_frame.grid(row=0, column=1, padx=10, pady=8, sticky=tk.W)
    
    forecast_entry = ttk.Entry(forecast_frame, width=10, textvariable=self.advanced_forecast_days_var)
    forecast_entry.pack(side=tk.LEFT)
    
    forecast_validation = ttk.Label(forecast_frame, text="", foreground=self.COLORS["danger"], style="TLabel")
    forecast_validation.pack(side=tk.LEFT, padx=5)
    
    # Add validation to forecast days entry
    def validate_forecast_days(*args):
        try:
            value = int(self.advanced_forecast_days_var.get())
            if value < 1:
                forecast_validation.configure(text="Must be > 0")
                return False
            elif value > 365:
                forecast_validation.configure(text="Max 365 days")
                return False
            else:
                forecast_validation.configure(text="")
                return True
        except ValueError:
            forecast_validation.configure(text="Enter a number")
            return False
    
    # Register the validation function
    self.advanced_forecast_days_var.trace("w", validate_forecast_days)
    
    # Volatility Parameters Section
    volatility_card = ttk.LabelFrame(right_col, text="Volatility Parameters", padding=15, style="TLabelframe")
    volatility_card.pack(fill=tk.X, pady=10)
    
    # Description
    vol_desc = ttk.Label(volatility_card, 
                       text="Adjust volatility parameters to account for different market conditions and risk factors.",
                       wraplength=400, style="TLabel")
    vol_desc.pack(anchor=tk.W, pady=(0, 10))
    
    # Volatility grid
    vol_grid = ttk.Frame(volatility_card, style="TFrame")
    vol_grid.pack(fill=tk.X, pady=5)
    
    # Implied Volatility Range
    ttk.Label(vol_grid, text="Min Volatility (%):", style="TLabel").grid(
        row=0, column=0, padx=10, pady=8, sticky=tk.W)
    min_vol_entry = ttk.Entry(vol_grid, width=10, textvariable=self.implied_volatility_low_var)
    min_vol_entry.grid(row=0, column=1, padx=10, pady=8, sticky=tk.W)
    
    ttk.Label(vol_grid, text="Max Volatility (%):", style="TLabel").grid(
        row=1, column=0, padx=10, pady=8, sticky=tk.W)
    max_vol_entry = ttk.Entry(vol_grid, width=10, textvariable=self.implied_volatility_high_var)
    max_vol_entry.grid(row=1, column=1, padx=10, pady=8, sticky=tk.W)
    
    # Volatility guidelines card with professional formatting
    guidelines_frame = ttk.LabelFrame(vol_grid, text="Market Condition Guidelines", padding=10)
    guidelines_frame.grid(row=0, column=2, rowspan=3, padx=10, pady=5, sticky=tk.W)
    
    # Guidelines content with bullet points
    guidelines_text = tk.Text(guidelines_frame, wrap=tk.WORD, height=6, width=30,
                           font=('Helvetica', 9), bd=0, highlightthickness=0,
                           background=self.COLORS["light"])
    guidelines_text.pack(fill=tk.BOTH, expand=True)
    guidelines_text.tag_configure("blue_text", foreground=self.COLORS["primary"])  # Use primary blue (same as title)
    guidelines_text.insert(tk.END, """• 15-20%: Normal market conditions
• 20-25%: Increased uncertainty
• 25-30%: High political risk
• 30-40%: Supply disruption risk
• 40-50%: Market crisis scenario""", "blue_text")
    guidelines_text.config(state=tk.DISABLED)  # Make read-only
    
    # Action Buttons section with modern styling
    action_card = ttk.LabelFrame(right_col, text="Analysis Actions", padding=15, style="TLabelframe")
    action_card.pack(fill=tk.X, pady=10)
    
    # Buttons with enhanced styling and clear purposes
    run_button = ttk.Button(action_card, text="Run VaR Simulation",
                          command=self.run_advanced_var_simulation, style="Primary.TButton")
    run_button.pack(fill=tk.X, pady=5)
    
    # Quick action buttons row
    quick_buttons = ttk.Frame(action_card, style="TFrame")
    quick_buttons.pack(fill=tk.X, pady=5)
    
    export_button = ttk.Button(quick_buttons, text="Export to Excel",
                             command=self.export_advanced_results, style="Primary.TButton")
    export_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
    
    reset_button = ttk.Button(quick_buttons, text="Reset Form",
                            command=self.reset_advanced_form, style="Primary.TButton")
    reset_button.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
    
    # Add copyright notice
    self.add_copyright_to_frame(frame)

def setup_results_view(self):
    """Setup the Results tab to show VaR simulation results and Seasonal Contracts"""
    frame = self.results_frame
    
    # Add explanation text
    header_label = ttk.Label(frame, 
                            text="VaR Simulation Results",
                            font=('Helvetica', 10, 'bold'))
    header_label.pack(fill=tk.X, pady=10)
    
    # Create a simple control panel for export
    control_panel = ttk.Frame(frame)
    control_panel.pack(fill=tk.X, pady=5)
    
    # Create extended values list with all possible contracts and custom price
    result_values = ["D1", "M1", "S1", "Custom"] + SEASONAL_CONTRACTS
    
    # Contract selection
    contract_frame = ttk.Frame(control_panel)
    contract_frame.pack(side=tk.LEFT, padx=10)
    
    ttk.Label(contract_frame, text="Select Contract:").pack(side=tk.LEFT, padx=5)
    contract_combo = ttk.Combobox(contract_frame, 
                                values=result_values,
                                width=15, textvariable=self.result_contract_var)
    contract_combo.pack(side=tk.LEFT)
    
    # Confidence level selection
    confidence_frame = ttk.Frame(control_panel)
    confidence_frame.pack(side=tk.LEFT, padx=10)
    
    ttk.Label(confidence_frame, text="Confidence Level:").pack(side=tk.LEFT, padx=5)
    confidence_combo = ttk.Combobox(confidence_frame, 
                                  values=["90%", "95%", "99%"],
                                  width=10, textvariable=self.result_confidence_var)
    confidence_combo.pack(side=tk.LEFT)
    
    # Action buttons
    button_frame = ttk.Frame(control_panel)
    button_frame.pack(side=tk.RIGHT, padx=10)
    
    export_button = ttk.Button(button_frame, text="Export to Excel",
                            command=self.export_results_tab, style="Primary.TButton")
    export_button.pack(side=tk.LEFT, padx=5)
    
    # Create a text area for displaying VaR results with monospaced font for proper table alignment
    self.results_text = tk.Text(frame, wrap=tk.NONE, height=35, font=('Courier', 12))
    self.results_text.pack(fill=tk.BOTH, expand=True, pady=5)
    
    # Add horizontal scrollbar for wide tables
    h_scrollbar = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=self.results_text.xview)
    h_scrollbar.pack(fill=tk.X, side=tk.BOTTOM, before=self.results_text)
    self.results_text.config(xscrollcommand=h_scrollbar.set)
    
    # Add scrollbar for the text area
    scrollbar = ttk.Scrollbar(self.results_text, command=self.results_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    self.results_text.config(yscrollcommand=scrollbar.set)
    
    # Refresh button
    refresh_button = ttk.Button(frame, text="Refresh Results",
                              command=self.update_results_tab, style="Primary.TButton")
    refresh_button.pack(pady=10)
    
    # Add copyright notice
    self.add_copyright_to_frame(frame)

def setup_visualization_view(self):
    """Setup the Visualisation tab for interactive visualization of VaR calculations"""
    frame = self.visualization_frame
    
    # Add explanation text
    header_label = ttk.Label(frame, 
                           text="VaR Visualisation Tools",
                           font=('Helvetica', 10, 'bold'))
    header_label.pack(fill=tk.X, pady=10)
    
    # Create controls panel 
    controls_frame = ttk.LabelFrame(frame, text="Visualization Controls", padding=10)
    controls_frame.pack(fill=tk.X, pady=5)
    
    # Create a grid for controls
    control_grid = ttk.Frame(controls_frame)
    control_grid.pack(fill=tk.X, pady=5)
    
    # Row 1: Contract selection and visualization type
    ttk.Label(control_grid, text="Contract:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
    
    # Create extended values list with all possible contracts and custom price
    vis_values = ["D1", "M1", "S1", "Custom"] + SEASONAL_CONTRACTS
    
    # Variables for visualization controls
    self.vis_contract_var = tk.StringVar(value="D1")
    self.vis_type_var = tk.StringVar(value="Histogram")
    self.vis_confidence_var = tk.StringVar(value="95%")
    self.vis_compare_var = tk.BooleanVar(value=False)
    self.vis_second_contract_var = tk.StringVar(value="M1")
    
    # Contract dropdown
    contract_combo = ttk.Combobox(control_grid, 
                                values=vis_values,
                                width=15, textvariable=self.vis_contract_var)
    contract_combo.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W)
    
    ttk.Label(control_grid, text="Visualization:").grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
    vis_type_combo = ttk.Combobox(control_grid, 
                               values=["Histogram", "Price Path", "Box Plot", "VaR Comparison"],
                               width=15, textvariable=self.vis_type_var)
    vis_type_combo.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
    
    # Row 2: Confidence level and comparison options
    ttk.Label(control_grid, text="Confidence:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
    confidence_combo = ttk.Combobox(control_grid, 
                                 values=["90%", "95%", "99%"],
                                 width=15, textvariable=self.vis_confidence_var)
    confidence_combo.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
    
    # Compare checkbox and second contract dropdown
    compare_check = ttk.Checkbutton(control_grid, text="Compare with:", 
                                  variable=self.vis_compare_var)
    compare_check.grid(row=1, column=2, padx=5, pady=5, sticky=tk.W)
    
    second_contract_combo = ttk.Combobox(control_grid, 
                                      values=vis_values,
                                      width=15, textvariable=self.vis_second_contract_var)
    second_contract_combo.grid(row=1, column=3, padx=5, pady=5, sticky=tk.W)
    
    # Row 3: Action buttons
    buttons_frame = ttk.Frame(control_grid)
    buttons_frame.grid(row=2, column=0, columnspan=4, pady=10)
    
    visualize_button = ttk.Button(buttons_frame, text="Generate Visualization",
                               command=self.generate_visualization, style="Primary.TButton")
    visualize_button.pack(side=tk.LEFT, padx=5)
    
    export_button = ttk.Button(buttons_frame, text="Export Chart",
                            command=self.export_visualization, style="Primary.TButton")
    export_button.pack(side=tk.LEFT, padx=5)
    
    # Create a frame for visualization
    self.visualization_container = ttk.LabelFrame(frame, text="Visualization", padding=10)
    self.visualization_container.pack(fill=tk.BOTH, expand=True, pady=10)
    
    # Add placeholder text
    placeholder_label = ttk.Label(self.visualization_container, 
                               text="Select parameters and click 'Generate Visualization' to create interactive charts",
                               font=('Helvetica', 9, 'italic'))
    placeholder_label.pack(pady=50)
    
    # Add copyright notice
    self.add_copyright_to_frame(frame)

# Removed seasonal results view

# Add core functionality methods
def pull_prices(self):
    """Pull latest prices for Advanced Mode"""
    try:
        if not self.current_excel_file:
            messagebox.showwarning(
                "No Data File", 
                "No data file is loaded. Please upload an Excel file in the Data Management tab."
            )
            return
            
        contract = self.contract_var.get()

        # Define the sheet names
        da_sheet = f"{contract}_da"  # Day-ahead
        mc_sheet = f"{contract}_mc"  # Month-ahead
        sc_sheet = f"{contract}_sc"  # Season-ahead
        s_sheet = f"{contract}_s"    # Seasonal futures
        
        # Update the seasonal info label
        self.seasonal_info_label.config(text=f"Retrieving prices for {contract.upper()} seasonal contracts for selected date...")

        # Get prices for basic contracts
        d1_price = get_latest_price(self.current_excel_file, da_sheet)
        m1_price = get_latest_price(self.current_excel_file, mc_sheet)
        s1_price = get_latest_price(self.current_excel_file, sc_sheet)
        
        # Update the UI with basic contracts
        if d1_price is not None:
            self.d1_price_var.set(f"{d1_price:.2f}")

        if m1_price is not None:
            self.m1_price_var.set(f"{m1_price:.2f}")

        if s1_price is not None:
            self.s1_price_var.set(f"{s1_price:.2f}")
        
        # Clear the selected seasons list
        self.selected_seasons = []
        
        # Get and update all seasonal prices
        seasonal_prices_found = 0
        for season in SEASONAL_CONTRACTS:
            # Get the price for this season
            season_price = get_latest_price(self.current_excel_file, s_sheet, season)
            
            # Store the price in our dictionary and update the UI
            if season_price is not None:
                # Store in the prices dictionary
                season_key = f"{contract}_{season}"
                self.seasonal_prices[season_key] = season_price
                
                # Update the corresponding StringVar
                if season in self.seasonal_price_vars:
                    self.seasonal_price_vars[season].set(f"{season_price:.2f}")
                    seasonal_prices_found += 1
                    
                    # Add this season to selected seasons if price is found
                    self.selected_seasons.append(season)
            else:
                # Clear the field if no price is found
                if season in self.seasonal_price_vars:
                    self.seasonal_price_vars[season].set("")
            
        # Provide feedback with the number of seasonal prices found
        if seasonal_prices_found > 0:
            self.seasonal_info_label.config(
                text=f"Successfully retrieved {seasonal_prices_found} seasonal prices for {contract.upper()}"
            )
            messagebox.showinfo(
                "Price Update", 
                f"Prices updated successfully. Found {seasonal_prices_found} seasonal prices for {contract.upper()}."
            )
        else:
            self.seasonal_info_label.config(
                text=f"No seasonal prices found for {contract.upper()}. Check your data source."
            )
            messagebox.showwarning(
                "Price Update", 
                f"Basic prices updated, but no seasonal prices found for {contract.upper()}."
            )

    except Exception as e:
        messagebox.showerror("Price Retrieval Error", f"Could not retrieve prices: {e}")

def run_advanced_var_simulation(self):
    """Run the Advanced Mode simulation"""
    try:
        if not self.current_excel_file:
            messagebox.showwarning(
                "No Data File", 
                "No data file is loaded. Please upload an Excel file in the Data Management tab."
            )
            return
            
        # Get parameters
        forecast_days = self.advanced_forecast_days_var.get()
        contract = self.contract_var.get()
        selected_date = self.date_entry.get_date()
        
        # Get implied volatility inputs and validate
        try:
            implied_vol_low = float(self.implied_volatility_low_var.get()) / 100.0
            implied_vol_high = float(self.implied_volatility_high_var.get()) / 100.0
            
            # Validate ranges
            if implied_vol_low < 0.01 or implied_vol_low > 1.0:
                messagebox.showwarning("Volatility Warning", "Min volatility should be between 1% and 100%")
                implied_vol_low = max(0.01, min(1.0, implied_vol_low))
                self.implied_volatility_low_var.set(str(implied_vol_low * 100))
                
            if implied_vol_high < 0.01 or implied_vol_high > 1.0:
                messagebox.showwarning("Volatility Warning", "Max volatility should be between 1% and 100%")
                implied_vol_high = max(0.01, min(1.0, implied_vol_high))
                self.implied_volatility_high_var.set(str(implied_vol_high * 100))
                
            if implied_vol_low > implied_vol_high:
                messagebox.showwarning("Volatility Warning", "Min volatility should be less than or equal to max volatility")
                implied_vol_low, implied_vol_high = implied_vol_high, implied_vol_low
                self.implied_volatility_low_var.set(str(implied_vol_low * 100))
                self.implied_volatility_high_var.set(str(implied_vol_high * 100))
                
        except ValueError:
            messagebox.showwarning("Volatility Warning", "Invalid volatility values, using defaults (15% and 20%)")
            implied_vol_low = 0.15
            implied_vol_high = 0.20
            self.implied_volatility_low_var.set("15")
            self.implied_volatility_high_var.set("20")

        # Get contract prices
        contract_prices = {}

        # Validate price inputs
        if self.d1_price_var.get():
            try:
                contract_prices["D1"] = float(self.d1_price_var.get())
            except ValueError:
                pass

        if self.m1_price_var.get():
            try:
                contract_prices["M1"] = float(self.m1_price_var.get())
            except ValueError:
                pass

        if self.s1_price_var.get():
            try:
                contract_prices["S1"] = float(self.s1_price_var.get())
            except ValueError:
                pass
                
        # Add all seasonal contracts that have prices
        for season in SEASONAL_CONTRACTS:
            if season in self.seasonal_price_vars and self.seasonal_price_vars[season].get():
                try:
                    contract_prices[season] = float(self.seasonal_price_vars[season].get())
                except ValueError:
                    pass
                    
        # Add custom price if provided
        if self.custom_price_var.get():
            try:
                contract_prices["Custom"] = float(self.custom_price_var.get())
            except ValueError:
                pass

        if not contract_prices:
            messagebox.showwarning("Simulation Warning", "No valid contract prices available")
            return

        # Reset previous results
        self.last_advanced_var_results = {
            "D1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None,
                   "full_simulations": None},
            "M1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None,
                   "full_simulations": None},
            "S1": {"current_price": None, "simulated_prices": [],
                   "var_90_lower": None, "var_90_upper": None,
                   "var_95_lower": None, "var_95_upper": None,
                   "var_99_lower": None, "var_99_upper": None,
                   "full_simulations": None},
            "date": selected_date.strftime("%Y-%m-%d"),
            "forecast_days": forecast_days,
            "implied_vol_low": implied_vol_low,
            "implied_vol_high": implied_vol_high
        }

        # Map contract types to sheet suffixes
        contract_sheet_map = {
            "D1": "da",  # Day-ahead
            "M1": "mc",  # Month-ahead
            "S1": "sc"   # Season-ahead
        }
        
        # Add results containers for all seasonal contracts with prices
        for season in SEASONAL_CONTRACTS:
            if season in contract_prices:
                # Add a placeholder for this season in the results
                self.last_advanced_var_results[season] = {
                    "current_price": None, 
                    "simulated_prices": [],
                    "var_90_lower": None, "var_90_upper": None,
                    "var_95_lower": None, "var_95_upper": None,
                    "var_99_lower": None, "var_99_upper": None,
                    "full_simulations": None
                }
                
        # Add result container for custom price if provided
        if "Custom" in contract_prices:
            self.last_advanced_var_results["Custom"] = {
                "current_price": None, 
                "simulated_prices": [],
                "var_90_lower": None, "var_90_upper": None,
                "var_95_lower": None, "var_95_upper": None,
                "var_99_lower": None, "var_99_upper": None,
                "full_simulations": None
            }

        # Run simulations for each contract
        for contract_type, initial_price in contract_prices.items():
            # Determine the correct sheet_name based on contract type
            if contract_type in ["D1", "M1", "S1"]:
                # Standard contract types
                sheet_suffix = contract_sheet_map.get(contract_type)
                if not sheet_suffix:
                    continue
                    
                sheet_name = f"{contract}_{sheet_suffix}"
            elif contract_type in SEASONAL_CONTRACTS:
                # This is a seasonal contract - use the seasonal sheet
                sheet_name = f"{contract}_s"
            elif contract_type == "Custom":
                # For custom price, use the default timeframe sheet
                timeframe = self.timeframe_var.get()
                sheet_name = f"{contract}_{timeframe}"
            else:
                # Unknown contract type
                continue
            
            # Try to load historical data
            log_returns = self.log_returns.get(sheet_name)
            
            if log_returns is None:
                print(f"No historical data available for {sheet_name}")
                # Load it now if needed
                log_returns = load_historical_data(excel_file=self.current_excel_file, sheet_name=sheet_name)
                if log_returns is not None:
                    self.log_returns[sheet_name] = log_returns
                else:
                    # Still no data available
                    print(f"Failed to load historical data for {sheet_name}")
                    continue

            # Calculate statistical parameters
            mean_log_return = np.mean(log_returns)
            std_log_return = np.std(log_returns)
            
            # Apply implied volatility adjustment based on user inputs
            # We'll adjust the volatility (standard deviation) by a factor between the low and high values
            vol_adjustment_range = np.linspace(implied_vol_low, implied_vol_high, NUM_SIMULATIONS)

            # Run simulation
            np.random.seed(42 + ord(contract_type[0]))  # Different seed for each contract
            simulations = np.zeros((NUM_SIMULATIONS, forecast_days))

            for i in range(NUM_SIMULATIONS):
                # Get volatility factor for this simulation
                vol_factor = vol_adjustment_range[i]
                
                # Apply volatility adjustment to this path
                adjusted_std = std_log_return * vol_factor
                
                random_log_returns = np.random.normal(
                    loc=mean_log_return,
                    scale=adjusted_std,  # Apply the volatility adjustment
                    size=forecast_days
                )
                price_path = initial_price * np.exp(np.cumsum(random_log_returns))
                simulations[i, :] = price_path

            # Calculate VaR at different confidence levels
            final_prices = simulations[:, -1]

            self.last_advanced_var_results[contract_type] = {
                "current_price": initial_price,
                "simulated_prices": final_prices.tolist(),
                "var_90_lower": np.percentile(final_prices, 10),
                "var_90_upper": np.percentile(final_prices, 90),
                "var_95_lower": np.percentile(final_prices, 5),
                "var_95_upper": np.percentile(final_prices, 95),
                "var_99_lower": np.percentile(final_prices, 1),
                "var_99_upper": np.percentile(final_prices, 99),
                "full_simulations": simulations,  # Store full simulation paths
                "sheet_name": sheet_name
            }
        
        # Display information about seasonal contracts and custom price
        seasonal_contracts = [s for s in contract_prices.keys() if s in SEASONAL_CONTRACTS]
        has_custom_price = "Custom" in contract_prices
        
        if seasonal_contracts or has_custom_price:
            if seasonal_contracts:
                seasonal_info = f"Included {len(seasonal_contracts)} seasonal contracts in simulation"
                print(seasonal_info)
                for season in seasonal_contracts:
                    print(f"  - {season}: {contract_prices[season]:.2f}")
            
            if has_custom_price:
                print(f"Included custom price: {contract_prices['Custom']:.2f}")
                
            # Seasonal tab has been removed
            pass

        messagebox.showinfo("Simulation Complete", "VaR simulation completed successfully.")
        self.update_results_tab()

    except ValueError as e:
        messagebox.showerror("Input Error", f"Invalid input: {e}")
    except Exception as e:
        messagebox.showerror("Simulation Error", f"An unexpected error occurred: {e}")

def update_results_tab(self):
    """Update the Results tab with VaR simulation results and Seasonal Contracts"""
    # Clear the text area
    self.results_text.config(state=tk.NORMAL)
    self.results_text.delete(1.0, tk.END)

    # Initialize content
    content = "=== VaR Simulation Results ===\n\n"
    
    adv_res = self.last_advanced_var_results
    if adv_res.get('date'):
        content += f"Analysis Date: {adv_res['date']}\n"
        content += f"Forecast Days: {adv_res.get('forecast_days', 10)}\n"
        
        # Add volatility settings if available
        if 'implied_vol_low' in adv_res and 'implied_vol_high' in adv_res:
            content += f"Implied Volatility Range: {adv_res['implied_vol_low']*100:.1f}% - {adv_res['implied_vol_high']*100:.1f}%\n\n"
        else:
            content += "\n"
    
    # Define contract types to display in main results (D1, M1, S1, Custom)
    main_contracts = ["D1", "M1", "S1", "Custom"]
    has_advanced_results = False
    
    # Check if we have any results
    for contract_type in main_contracts:
        if contract_type in adv_res and adv_res[contract_type]["current_price"] is not None:
            has_advanced_results = True
            break
    
    # If no results, show placeholder
    if not has_advanced_results:
        content += "\nNo simulation results available. Run a simulation to see results here.\n"
    else:
        # Add the VaR simulation results table with better formatting
        header = "{:<10} | {:<13} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15} | {:<15}\n".format(
            "Contract", "Current Price", "VaR (90% lower)", "VaR (90% upper)", 
            "VaR (95% lower)", "VaR (95% upper)", "VaR (99% lower)", "VaR (99% upper)")
        
        separator = "-" * 10 + " | " + "-" * 13 + " | " + "-" * 15 + " | " + "-" * 15 + " | " + \
                   "-" * 15 + " | " + "-" * 15 + " | " + "-" * 15 + " | " + "-" * 15 + "\n"
        
        content += header
        content += separator
        
        # Show main contracts (D1, M1, S1, Custom)
        for contract_type in main_contracts:
            if contract_type in adv_res and adv_res[contract_type]["current_price"] is not None:
                res = adv_res[contract_type]
                
                # Format the contract name
                if contract_type in ["D1", "M1", "S1"]:
                    display_name = contract_type
                else:
                    display_name = "Custom"
                
                # Format each row with proper spacing
                row = "{:<10} | {:<13.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f}\n".format(
                    display_name, res['current_price'], 
                    res['var_90_lower'], res['var_90_upper'],
                    res['var_95_lower'], res['var_95_upper'],
                    res['var_99_lower'], res['var_99_upper']
                )
                content += row
        
        # Add seasonal contracts with results
        seasonal_contracts = [k for k in adv_res.keys() if k in SEASONAL_CONTRACTS and adv_res[k]["current_price"] is not None]
        
        if seasonal_contracts:
            content += "\n=== Seasonal Contracts Results ===\n\n"
            content += header
            content += separator
            
            for contract_type in seasonal_contracts:
                res = adv_res[contract_type]
                
                # Format each row with proper spacing
                row = "{:<10} | {:<13.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f} | {:<15.2f}\n".format(
                    contract_type, res['current_price'], 
                    res['var_90_lower'], res['var_90_upper'],
                    res['var_95_lower'], res['var_95_upper'],
                    res['var_99_lower'], res['var_99_upper']
                )
                content += row
    
    # Use a fixed-width font for better table display
    self.results_text.insert(tk.END, content)
    
    # Configure tag for monospaced font to ensure proper column alignment
    self.results_text.tag_configure("monospace", font=("Courier", 12))
    self.results_text.tag_add("monospace", "1.0", "end")
    
    self.results_text.config(state=tk.DISABLED)

def visualize_selected_results(self):
    """Update the visualization based on user selections in the Results tab"""
    self.update_results_visualization()

def update_results_visualization(self):
    """Update the visualization in Results tab based on current selections"""
    # This method is no longer used since visualization is now in its own tab
    pass

# Removed seasonal results tab function

def export_advanced_results(self, contract_type=None, confidence=90):
    """Export Advanced Mode simulation results to Excel"""
    # Check if we have any results
    if not any(self.last_advanced_var_results[k]['current_price'] for k in ["D1", "M1", "S1"]):
        messagebox.showwarning("Export Error", "No Advanced Mode simulation results available to export")
        return
        
    # Ask for save location
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Save Advanced Mode Results As"
    )
    
    if not filename:
        return  # User cancelled
        
    # Export the results
    if export_to_excel(self.last_advanced_var_results, filename):
        messagebox.showinfo("Export Success", f"Results exported to {filename}")
    else:
        messagebox.showerror("Export Error", "Failed to export results to Excel")

def export_results_tab(self):
    """Export current results from Results tab to Excel"""
    # Get selected contract
    contract = self.result_contract_var.get()
    self.export_advanced_results(contract_type=contract)
    
# Removed seasonal results export function

# Define visualization methods
def generate_visualization(self):
    """Generate interactive visualization based on selected parameters"""
    # Clear existing visualization container
    for widget in self.visualization_container.winfo_children():
        widget.destroy()
        
    # Get selected parameters
    contract = self.vis_contract_var.get()
    vis_type = self.vis_type_var.get()
    confidence_str = self.vis_confidence_var.get()
    confidence = int(confidence_str.replace('%', ''))
    compare = self.vis_compare_var.get()
    second_contract = self.vis_second_contract_var.get() if compare else None
    
    # Check if we have simulation results for the selected contract
    adv_res = self.last_advanced_var_results
    if contract not in adv_res or adv_res[contract]["current_price"] is None:
        ttk.Label(self.visualization_container, 
                text=f"No simulation results available for {contract}. Run a simulation first.",
                font=('Helvetica', 10)).pack(pady=50)
        return
    
    # If comparing, check if second contract has results
    if compare and (second_contract not in adv_res or adv_res[second_contract]["current_price"] is None):
        ttk.Label(self.visualization_container, 
                text=f"No simulation results available for comparison contract {second_contract}.",
                font=('Helvetica', 10)).pack(pady=50)
        return
    
    # Create figure for visualization with good size for displaying
    fig = Figure(figsize=(10, 6), dpi=100)
    
    # Different visualizations based on selected type
    if vis_type == "Histogram":
        self.create_histogram_visualization(fig, contract, confidence, compare, second_contract)
    elif vis_type == "Price Path":
        self.create_price_path_visualization(fig, contract, confidence, compare, second_contract)
    elif vis_type == "Box Plot":
        self.create_box_plot_visualization(fig, contract, compare, second_contract)
    elif vis_type == "VaR Comparison":
        self.create_var_comparison_visualization(fig, contract, confidence, compare, second_contract)
    
    # Add the visualization to the container
    canvas = FigureCanvasTkAgg(fig, master=self.visualization_container)
    canvas.draw()
    toolbar_frame = ttk.Frame(self.visualization_container)
    toolbar_frame.pack(fill=tk.X, pady=5)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    # Add note about interactivity
    note_label = ttk.Label(self.visualization_container, 
                         text="Tip: You can zoom, pan, and save these visualizations using the toolbar.",
                         font=('Helvetica', 9, 'italic'))
    note_label.pack(pady=5)
    
    # Add copyright to visualization container
    copyright_label = ttk.Label(
        self.visualization_container, 
        text=self.copyright_text,
        foreground="#AAAAAA",
        font=('Helvetica', 8),
        anchor=tk.E
    )
    copyright_label.pack(side=tk.BOTTOM, anchor=tk.SE, padx=5, pady=2)

def create_histogram_visualization(self, fig, contract, confidence, compare, second_contract=None):
    """Create histogram visualization of price distribution"""
    # Get results data
    adv_res = self.last_advanced_var_results
    contract_res = adv_res[contract]
    
    # Set up subplot
    ax = fig.add_subplot(111)
    
    # Get simulated prices
    sim_prices = contract_res['simulated_prices']
    current_price = contract_res['current_price']
    
    # Get appropriate VaR levels based on confidence
    if confidence == 90:
        var_lower = contract_res['var_90_lower']
        var_upper = contract_res['var_90_upper']
    elif confidence == 95:
        var_lower = contract_res['var_95_lower']
        var_upper = contract_res['var_95_upper']
    else:  # 99%
        var_lower = contract_res['var_99_lower']
        var_upper = contract_res['var_99_upper']
    
    # Create main histogram
    ax.hist(sim_prices, bins=40, alpha=0.7, color='skyblue', edgecolor='black', 
         label=f'{contract} (Current: {current_price:.2f})')
    
    # Add comparison histogram if requested
    if compare and second_contract:
        second_res = adv_res[second_contract]
        second_prices = second_res['simulated_prices']
        second_current = second_res['current_price']
        
        # Scale prices to make comparison more visually meaningful
        if abs(current_price - second_current) / max(current_price, second_current) > 0.2:
            # Prices differ by more than 20% - normalize for better visualization
            scale_factor = current_price / second_current
            scaled_prices = [p * scale_factor for p in second_prices]
            ax.hist(scaled_prices, bins=40, alpha=0.5, color='lightgreen', edgecolor='black',
                 label=f'{second_contract} (scaled, Current: {second_current:.2f})')
            
            # Add note about scaling
            ax.text(0.02, 0.02, f"Note: {second_contract} prices scaled by factor {scale_factor:.2f} for comparison",
                 transform=ax.transAxes, fontsize=8, alpha=0.7)
        else:
            # Prices are similar enough for direct comparison
            ax.hist(second_prices, bins=40, alpha=0.5, color='lightgreen', edgecolor='black',
                 label=f'{second_contract} (Current: {second_current:.2f})')
    
    # Add reference lines
    ax.axvline(current_price, color='blue', linestyle=':', linewidth=2,
             label=f'Current Price: {current_price:.2f}')
    ax.axvline(var_lower, color='red', linestyle='--', linewidth=2,
             label=f'{confidence}% VaR Lower: {var_lower:.2f}')
    ax.axvline(var_upper, color='red', linestyle='--', linewidth=2,
             label=f'{confidence}% VaR Upper: {var_upper:.2f}')
    
    # Add labels and title
    contract_display = contract if contract in ["D1", "M1", "S1"] else contract  # Format contract name
    ax.set_title(f"Distribution of Simulated Prices for {contract_display} - {confidence}% Confidence")
    ax.set_xlabel("Price")
    ax.set_ylabel("Frequency")
    
    # Add volatility info if available
    if 'implied_vol_low' in adv_res and 'implied_vol_high' in adv_res:
        vol_info = f"(Volatility: {adv_res['implied_vol_low']*100:.1f}%-{adv_res['implied_vol_high']*100:.1f}%)"
        ax.set_title(f"Distribution of Simulated Prices for {contract_display} - {confidence}% Confidence\n{vol_info}")
    
    # Add legend
    ax.legend(loc='upper right')
    
    # Add grid for readability
    ax.grid(True, alpha=0.3)
    
    # Ensure layout looks good
    fig.tight_layout()

def create_price_path_visualization(self, fig, contract, confidence, compare, second_contract=None):
    """Create price path visualization showing possible price paths over time"""
    # Get results data
    adv_res = self.last_advanced_var_results
    contract_res = adv_res[contract]
    
    # Check if we have full simulations data (not just final prices)
    if 'full_simulations' not in contract_res or contract_res['full_simulations'] is None:
        ax = fig.add_subplot(111)
        ax.text(0.5, 0.5, "Price path visualization requires full simulation data.\nRun a new simulation to view price paths.",
              ha='center', va='center', fontsize=12)
        fig.tight_layout()
        return
    
    # Set up subplot
    ax = fig.add_subplot(111)
    
    # Get simulation data
    simulations = contract_res['full_simulations']
    forecast_days = adv_res.get('forecast_days', 10)
    current_price = contract_res['current_price']
    
    # Create time points
    time_points = np.arange(forecast_days)
    
    # Plot a sample of paths (too many would be unreadable)
    num_sample_paths = min(50, simulations.shape[0])
    sample_indices = np.random.choice(simulations.shape[0], num_sample_paths, replace=False)
    
    # Plot each sampled path with slight transparency
    for idx in sample_indices:
        ax.plot(time_points, simulations[idx], color='skyblue', alpha=0.3, linewidth=0.8)
    
    # Plot mean path
    mean_path = np.mean(simulations, axis=0)
    ax.plot(time_points, mean_path, color='blue', linewidth=2, label='Mean Path')
    
    # Get appropriate VaR levels based on confidence
    if confidence == 90:
        percentile_low = 10
        percentile_high = 90
    elif confidence == 95:
        percentile_low = 5
        percentile_high = 95
    else:  # 99%
        percentile_low = 1
        percentile_high = 99
        
    # Calculate confidence intervals for each time point
    lower_bound = np.percentile(simulations, percentile_low, axis=0)
    upper_bound = np.percentile(simulations, percentile_high, axis=0)
    
    # Plot confidence interval
    ax.fill_between(time_points, lower_bound, upper_bound, color='lightblue', alpha=0.5,
                  label=f'{confidence}% Confidence Interval')
    
    # Add comparison if requested
    if compare and second_contract:
        second_res = adv_res[second_contract]
        
        if 'full_simulations' in second_res and second_res['full_simulations'] is not None:
            second_simulations = second_res['full_simulations']
            second_current = second_res['current_price']
            
            # Scale if prices differ significantly
            if abs(current_price - second_current) / max(current_price, second_current) > 0.2:
                # Scale factor for comparison
                scale_factor = current_price / second_current
                
                # Sample and plot scaled paths
                for idx in sample_indices[:25]:  # Reduce number to avoid clutter
                    ax.plot(time_points, second_simulations[idx] * scale_factor, 
                           color='lightgreen', alpha=0.3, linewidth=0.8)
                
                # Plot scaled mean path
                second_mean = np.mean(second_simulations, axis=0) * scale_factor
                ax.plot(time_points, second_mean, color='green', linewidth=2, 
                       label=f'{second_contract} Mean Path (scaled)')
                
                # Add note about scaling
                ax.text(0.02, 0.02, f"Note: {second_contract} prices scaled by factor {scale_factor:.2f} for comparison",
                       transform=ax.transAxes, fontsize=8, alpha=0.7)
            else:
                # Sample and plot direct paths
                for idx in sample_indices[:25]:  # Reduce number to avoid clutter
                    ax.plot(time_points, second_simulations[idx], color='lightgreen', alpha=0.3, linewidth=0.8)
                
                # Plot mean path
                second_mean = np.mean(second_simulations, axis=0)
                ax.plot(time_points, second_mean, color='green', linewidth=2, 
                       label=f'{second_contract} Mean Path')
    
    # Add labels and title
    contract_display = contract if contract in ["D1", "M1", "S1"] else contract
    ax.set_title(f"Simulated Price Paths for {contract_display} over {forecast_days} Days")
    ax.set_xlabel("Days Forward")
    ax.set_ylabel("Price")
    
    # Starting point
    ax.scatter([0], [current_price], color='blue', s=50, zorder=5, label=f'Starting Price: {current_price:.2f}')
    
    # Add grid and legend
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best')
    
    # Ensure layout looks good
    fig.tight_layout()

def create_box_plot_visualization(self, fig, contract, compare, second_contract=None):
    """Create box plot visualization of price distributions"""
    # Get results data
    adv_res = self.last_advanced_var_results
    
    # Set up subplot
    ax = fig.add_subplot(111)
    
    # Prepare data
    box_data = []
    labels = []
    
    # Add main contract data
    contract_res = adv_res[contract]
    sim_prices = contract_res['simulated_prices']
    box_data.append(sim_prices)
    labels.append(contract)
    
    # Add comparison contract if requested
    if compare and second_contract:
        second_res = adv_res[second_contract]
        second_prices = second_res['simulated_prices']
        box_data.append(second_prices)
        labels.append(second_contract)
    
    # Create box plot
    boxplot = ax.boxplot(box_data, patch_artist=True, labels=labels)
    
    # Customize box plot colors
    colors = ['skyblue', 'lightgreen', 'lightpink', 'lightyellow']
    for i, box in enumerate(boxplot['boxes']):
        box.set(facecolor=colors[i % len(colors)])
    
    # Add current prices as points
    x_positions = np.arange(1, len(labels) + 1)
    current_prices = []
    
    # Main contract
    current_prices.append(contract_res['current_price'])
    
    # Comparison contract
    if compare and second_contract:
        current_prices.append(second_res['current_price'])
    
    # Plot current prices
    ax.scatter(x_positions, current_prices, color='red', marker='o', s=50, zorder=5, label='Current Prices')
    
    # Label current prices
    for i, price in enumerate(current_prices):
        ax.annotate(f"{price:.2f}", (x_positions[i], price), 
                  xytext=(0, 10), textcoords='offset points',
                  ha='center', fontsize=9)
    
    # Add title and labels
    ax.set_title(f"Box Plot of Simulated Price Distributions")
    ax.set_xlabel("Contract")
    ax.set_ylabel("Price")
    
    # Add volatility info if available
    if 'implied_vol_low' in adv_res and 'implied_vol_high' in adv_res:
        vol_info = f"(Volatility: {adv_res['implied_vol_low']*100:.1f}%-{adv_res['implied_vol_high']*100:.1f}%)"
        ax.set_title(f"Box Plot of Simulated Price Distributions\n{vol_info}")
    
    # Add legend for current prices
    ax.legend(loc='best')
    
    # Add grid for readability
    ax.grid(True, alpha=0.3)
    
    # Add box plot interpretation guide
    explanation = (
        "Box Plot Guide:\n"
        "- Middle line: Median\n"
        "- Box: 25th-75th percentiles\n"
        "- Whiskers: Min/Max (excludes outliers)\n"
        "- Red dot: Current price"
    )
    ax.text(0.02, 0.02, explanation, transform=ax.transAxes, 
          fontsize=8, va='bottom', ha='left', bbox=dict(boxstyle='round', alpha=0.1))
    
    # Ensure layout looks good
    fig.tight_layout()

def create_var_comparison_visualization(self, fig, contract, confidence, compare, second_contract=None):
    """Create VaR comparison visualization across confidence levels"""
    # Get results data
    adv_res = self.last_advanced_var_results
    contract_res = adv_res[contract]
    
    # Set up subplot
    ax = fig.add_subplot(111)
    
    # Current price data
    current_price = contract_res['current_price']
    
    # Define confidence levels to show
    confidence_levels = [90, 95, 99]
    
    # Prepare data for the selected contract
    var_lowers = [contract_res[f'var_{cl}_lower'] for cl in confidence_levels]
    var_uppers = [contract_res[f'var_{cl}_upper'] for cl in confidence_levels]
    
    # Calculate price ranges and percentages
    price_ranges = [upper - lower for upper, lower in zip(var_uppers, var_lowers)]
    pct_ranges = [(r / current_price) * 100 if current_price > 0 else 0 for r in price_ranges]
    
    # X positions for the bars
    x_pos = np.arange(len(confidence_levels))
    bar_width = 0.35
    
    # Create bar chart for price ranges
    bars1 = ax.bar(x_pos - bar_width/2, price_ranges, bar_width, label=f'{contract} Price Range', 
                color='skyblue', edgecolor='black')
    
    # Add comparison if requested
    if compare and second_contract:
        second_res = adv_res[second_contract]
        second_current = second_res['current_price']
        
        # Calculate second contract data
        second_lowers = [second_res[f'var_{cl}_lower'] for cl in confidence_levels]
        second_uppers = [second_res[f'var_{cl}_upper'] for cl in confidence_levels]
        second_ranges = [upper - lower for upper, lower in zip(second_uppers, second_lowers)]
        
        # Create comparison bars
        bars2 = ax.bar(x_pos + bar_width/2, second_ranges, bar_width, 
                     label=f'{second_contract} Price Range', color='lightgreen', edgecolor='black')
    
    # Add labels and title
    ax.set_title(f"VaR Price Ranges at Different Confidence Levels")
    ax.set_xlabel("Confidence Level")
    ax.set_ylabel("Price Range")
    
    # Set x-axis labels
    ax.set_xticks(x_pos)
    ax.set_xticklabels([f"{cl}%" for cl in confidence_levels])
    
    # Add price range values as text labels on bars
    for i, v in enumerate(price_ranges):
        ax.text(i - bar_width/2, v + 0.1, f"{v:.2f}", ha='center', fontsize=9)
        
    if compare and second_contract:
        for i, v in enumerate(second_ranges):
            ax.text(i + bar_width/2, v + 0.1, f"{v:.2f}", ha='center', fontsize=9)
    
    # Add percentage axis on the right
    ax2 = ax.twinx()
    ax2.set_ylabel("% of Current Price")
    
    # Plot percentage lines for main contract
    ax2.plot(x_pos, pct_ranges, 'o-', color='blue', linewidth=2, label=f'{contract} % of Price')
    
    # Plot percentage lines for comparison contract if requested
    if compare and second_contract:
        second_pct = [(r / second_current) * 100 if second_current > 0 else 0 for r in second_ranges]
        ax2.plot(x_pos, second_pct, 'o-', color='green', linewidth=2, label=f'{second_contract} % of Price')
    
    # Add all legends
    handles1, labels1 = ax.get_legend_handles_labels()
    handles2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(handles1 + handles2, labels1 + labels2, loc='upper left')
    
    # Add current price info
    price_info = f"{contract} current price: {current_price:.2f}"
    if compare and second_contract:
        price_info += f" | {second_contract} current price: {second_current:.2f}"
    ax.text(0.5, 0.02, price_info, transform=ax.transAxes, ha='center', fontsize=9)
    
    # Add grid for readability
    ax.grid(True, alpha=0.3)
    
    # Ensure layout looks good
    fig.tight_layout()

def export_visualization(self):
    """Export the current visualization to a file"""
    # Check if we have a visualization to export
    if not hasattr(self, 'visualization_container') or not self.visualization_container.winfo_children():
        messagebox.showwarning("Export Warning", "No visualization available to export")
        return
    
    # Ask user for file name and location
    filetypes = [
        ("PNG files", "*.png"),
        ("PDF files", "*.pdf"),
        ("SVG files", "*.svg"),
        ("JPEG files", "*.jpg"),
        ("All files", "*.*")
    ]
    
    filename = filedialog.asksaveasfilename(
        title="Export Visualization",
        filetypes=filetypes,
        defaultextension=".png"
    )
    
    if not filename:
        return  # User cancelled
    
    try:
        # Find the FigureCanvasTkAgg in the visualization container
        canvas = None
        for widget in self.visualization_container.winfo_children():
            if isinstance(widget, FigureCanvasTkAgg):
                canvas = widget
                break
        
        if canvas:
            fig = canvas.figure
            fig.savefig(filename, dpi=300, bbox_inches='tight')
            messagebox.showinfo("Export Success", f"Visualization exported to {filename}")
        else:
            messagebox.showwarning("Export Warning", "No visualization canvas found to export")
    
    except Exception as e:
        messagebox.showerror("Export Error", f"Error exporting visualization: {e}")

# Add the methods to the MonteCarloVaRApp class
MonteCarloVaRApp.setup_advanced_mode = setup_advanced_mode
MonteCarloVaRApp.setup_results_view = setup_results_view
MonteCarloVaRApp.setup_visualization_view = setup_visualization_view
MonteCarloVaRApp.pull_prices = pull_prices
MonteCarloVaRApp.run_advanced_var_simulation = run_advanced_var_simulation
MonteCarloVaRApp.update_results_tab = update_results_tab
MonteCarloVaRApp.visualize_selected_results = visualize_selected_results
MonteCarloVaRApp.update_results_visualization = update_results_visualization
MonteCarloVaRApp.generate_visualization = generate_visualization
MonteCarloVaRApp.create_histogram_visualization = create_histogram_visualization
MonteCarloVaRApp.create_price_path_visualization = create_price_path_visualization
MonteCarloVaRApp.create_box_plot_visualization = create_box_plot_visualization
MonteCarloVaRApp.create_var_comparison_visualization = create_var_comparison_visualization
MonteCarloVaRApp.export_visualization = export_visualization
MonteCarloVaRApp.export_advanced_results = export_advanced_results
MonteCarloVaRApp.export_results_tab = export_results_tab

def main():
    """Entry point for the application"""
    try:
        app = MonteCarloVaRApp()
        
        # When the app starts for the first time, show a welcome message
        # and offer to create a desktop shortcut
        data_loaded = app.current_excel_file is not None
        welcome_message = (
            "Welcome to the Monte Carlo VaR Risk Tool!\n\n"
        )
        
        if not data_loaded:
            welcome_message += (
                "No data file was found. Please go to the 'Data Management' tab "
                "to upload an Excel file (lu_energy_prices_historic.xlsx) to begin."
            )
        else:
            welcome_message += (
                "Data file loaded successfully.\n"
                "You can update the data file at any time through the 'Data Management' tab."
            )
        
        welcome_message += "\n\nWould you like to create a desktop shortcut for easy access?"
        
        # Ask about creating a shortcut
        create_shortcut = messagebox.askyesno(
            "Welcome", 
            welcome_message
        )
        
        if create_shortcut:
            success, message = create_desktop_shortcut()
            if success:
                messagebox.showinfo("Shortcut Created", message)
            else:
                messagebox.showerror("Shortcut Error", message)
        
        app.mainloop()
    except Exception as e:
        print(f"An error occurred while running the application: {e}")


if __name__ == "__main__":
    main()